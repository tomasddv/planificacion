from __future__ import annotations

import io
import json
import os
import re
import shutil
import time
import unicodedata
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

try:
    import holidays
except ImportError:  # pragma: no cover - app keeps working without holiday package.
    holidays = None


APP_TITLE = "Venta diaria HL"
SMALL_DASH_URL = "https://planificacion-ifeevprb7is4zwjk6k5suo.streamlit.app/"
PROJECT_ROOT = Path(__file__).resolve().parent
DATA_DIR_CANDIDATES = [
    PROJECT_ROOT / "planificacion",
    PROJECT_ROOT / "data",
    Path(os.environ.get("DASHBOARDS_ROOT", "N:/Tomas/DASHBOARDS")) / "planificacion",
    Path("N:/tomas/dashboards/planificacion"),
    Path.home() / "Desktop" / "planificacion",
]
VALID_EXTENSIONS = {".txt", ".csv"}
CLIENT_EXTENSIONS = {".xlsx", ".xls", ".csv", ".txt"}
PLANNER_STORE_FILE_NAME = "planificador_diario_guardado.csv"
PLANNER_DATA_DIR = Path(os.environ.get("PLANNER_DATA_DIR", PROJECT_ROOT / ".planner_data"))
PLANNER_SHEET_NAME = "planificador_diario"
PLANNER_COLUMNS = ["fecha", "foco", "promotor", "planificado"]
WINDOWS = (7, 14, 21, 28)
EXACT_MONTH_LOOKBACKS = (1, 2, 3)
NORMALIZATION_VERSION = 4
CUSTOMER_CHANNEL_VERSION = 2
SEGMENT_VERSION = 1
CANAL_ORDER = ["K+T", "AUTOSERVICIO", "MAYORISTA", "REF", "NO"]
DIVISION_REPORT_ORDER = [
    "TOTAL CVZA",
    "TOTAL UNG",
    "CVZA CORE + VALUE",
    "CVZA CORE",
    "CVZA VALUE",
    "CVZA CORE +",
    "CVZA HE",
    "UNG SIN TOP",
    "UNG TOP",
    "AGUAS ECO",
    "VINO",
    "ADYACENCIAS",
]
REPORT_TOTAL_UNITS = {"CZA", "UNG", "AGUAS ECO", "VINO", "ADYACENCIAS"}
PLANNER_FOCUS_RULES = {
    "Foco 1 - Total Cervezas 2026": {
        "title": "TOTAL CZA",
        "caption": "CERVEZAS, GIFTPACK CERVEZAS, POP",
        "unidad_negocio": {"CZA"},
    },
    "Foco 2 - Above core 2026": {
        "title": "ABOVE CORE",
        "caption": "CERVEZAS, GIFTPACK CERVEZAS, SIDRAS, POP",
        "division_informe": {"CVZA HE", "CVZA CORE +"},
    },
    "Foco 3 - Total UNG 2026": {
        "title": "TOTAL UNG",
        "caption": "UNG",
        "unidad_negocio": {"UNG"},
    },
    "Foco 4 - Total Aguas 2026": {
        "title": "TOTAL AGUAS",
        "caption": "GASEOSAS, AGUAS, POP AGUAS",
        "division_informe": {"AGUAS ECO"},
    },
}
PLANNER_OBJECTIVE_ALIASES = {
    "TOTAL CZA": "Foco 1 - Total Cervezas 2026",
    "TOTAL CVZA": "Foco 1 - Total Cervezas 2026",
    "TOTAL CERVEZAS": "Foco 1 - Total Cervezas 2026",
    "ABOVE CORE": "Foco 2 - Above core 2026",
    "VOLUMEN ABOVE CORE": "Foco 2 - Above core 2026",
    "TOTAL UNG": "Foco 3 - Total UNG 2026",
    "UNG": "Foco 3 - Total UNG 2026",
    "AGUAS": "Foco 4 - Total Aguas 2026",
    "TOTAL AGUAS": "Foco 4 - Total Aguas 2026",
    "AGUAS ECO": "Foco 4 - Total Aguas 2026",
}
DEFAULT_OBJECTIVES_ROWS = [
    {"seccion": "", "item": "TOTAL CVZA", "OBJ VTAS": 3633.51},
    {"seccion": "", "item": "TOTAL UNG", "OBJ VTAS": 2923.65},
    {"seccion": "", "item": "CZA", "OBJ VTAS": 3633.51},
    {"seccion": "", "item": "UNG", "OBJ VTAS": 2923.65},
    {"seccion": "", "item": "CVZA HE", "OBJ VTAS": 952.22},
    {"seccion": "", "item": "CVZA CORE + VALUE", "OBJ VTAS": 2681.29},
    {"seccion": "", "item": "MIX PREMIUM", "OBJ VTAS": 26.2},
    {"seccion": "", "item": "AGUAS ECO", "OBJ VTAS": 240.84},
]
PROMOTER_MESA_MAP = {
    "NICASTRO LUCAS": "ismael bruno",
    "SIRI MARTIN": "ismael bruno",
    "MATIAS GARCIA": "ismael bruno",
    "GASTON FABRE": "ismael bruno",
    "NICOLAS POCHETINO": "ismael bruno",
    "VILLAGRA ENZO": "ismael bruno",
    "FEDERICO BISS": "anibal viti",
    "PABLO ALVAREZ": "casco hernan",
    "ALEXANDER ROJAS": "casco hernan",
    "FERNANDO FIELG": "casco hernan",
    "JUAN MANUEL GIMENEZ": "casco hernan",
    "MENDEZ CARLOS": "casco hernan",
    "MARIANO HERRERA": "casco hernan",
}
VENDOR_NAME_ALIASES = {
    "ENZO VILLAGRA": "VILLAGRA ENZO",
}


def secret_or_env(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name, "")
    except Exception:
        value = ""
    return str(value or os.environ.get(name, default)).strip()


def truthy(value: str) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "si", "sí", "y"}


def resolve_google_drive_folder(secret_name: str, folder_name: str, force_refresh: bool = False) -> Path | None:
    url = secret_or_env(secret_name)
    if not url:
        return None

    target = PROJECT_ROOT / ".cloud_data" / folder_name
    refresh = force_refresh or truthy(secret_or_env("FORCE_GDRIVE_REFRESH", "false"))
    has_files = target.exists() and any(target.iterdir())
    if has_files and not refresh:
        return target

    try:
        import gdown
    except ImportError:
        return None

    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_target = PROJECT_ROOT / ".cloud_data" / f"{folder_name}_tmp_{int(time.time())}"
    if tmp_target.exists():
        shutil.rmtree(tmp_target, ignore_errors=True)
    tmp_target.mkdir(parents=True, exist_ok=True)

    try:
        gdown.download_folder(url=url, output=str(tmp_target), quiet=True, use_cookies=False)
    except Exception:
        return target if target.exists() and any(target.iterdir()) else None

    if tmp_target.exists() and any(tmp_target.iterdir()):
        if target.exists():
            shutil.rmtree(target, ignore_errors=True)
        try:
            tmp_target.rename(target)
            return target
        except Exception:
            return tmp_target
    return target if target.exists() and any(target.iterdir()) else None


DEFAULT_DATA_DIR = (
    resolve_google_drive_folder("GOOGLE_DRIVE_PLANIFICACION_URL", "planificacion")
    or next((path for path in DATA_DIR_CANDIDATES if path.exists()), DATA_DIR_CANDIDATES[0])
)


def current_data_dir(force_refresh: bool = False) -> Path:
    return (
        resolve_google_drive_folder("GOOGLE_DRIVE_PLANIFICACION_URL", "planificacion", force_refresh=force_refresh)
        or next((path for path in DATA_DIR_CANDIDATES if path.exists()), DATA_DIR_CANDIDATES[0])
    )


@dataclass(frozen=True)
class SourceInfo:
    label: str
    path: str | None
    modified: str | None


def excel_col_to_index(letter: str) -> int:
    value = 0
    for char in letter.upper():
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value - 1


def page_setup() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon=":bar_chart:", layout="wide")
    st.markdown(
        """
        <style>
        :root {
            --bg: #f5f7fb;
            --ink: #101828;
            --muted: #667085;
            --blue: #1463ff;
            --cyan: #00a7c8;
            --green: #12b76a;
            --orange: #f79009;
            --red: #f04438;
            --violet: #7a5af8;
            --card: rgba(255,255,255,.92);
        }
        .stApp {
            background:
                radial-gradient(circle at 12% 8%, rgba(20, 99, 255, .13), transparent 28%),
                radial-gradient(circle at 88% 6%, rgba(18, 183, 106, .13), transparent 24%),
                linear-gradient(180deg, #f8fbff 0%, var(--bg) 42%, #eef3fb 100%);
            color: var(--ink);
        }
        .stApp, .stApp p, .stApp span, .stApp label, .stApp div {
            color: var(--ink);
        }
        .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0f172a 0%, #172554 100%);
        }
        [data-testid="stSidebar"] h1,
        [data-testid="stSidebar"] h2,
        [data-testid="stSidebar"] h3,
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] small,
        [data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
            color: #f8fafc !important;
        }
        [data-testid="stSidebar"] input,
        [data-testid="stSidebar"] textarea,
        [data-testid="stSidebar"] [data-baseweb="select"] *,
        [data-testid="stSidebar"] [data-baseweb="input"] * {
            color: #111827 !important;
        }
        [data-testid="stSidebar"] [data-baseweb="select"] > div,
        [data-testid="stSidebar"] [data-baseweb="input"],
        [data-testid="stSidebar"] [data-baseweb="base-input"],
        [data-testid="stSidebar"] .stDateInput input,
        [data-testid="stSidebar"] .stMultiSelect div,
        [data-testid="stSidebar"] .stSelectbox div {
            background-color: #ffffff !important;
            color: #111827 !important;
        }
        [data-testid="stSidebar"] [data-baseweb="select"] input,
        [data-testid="stSidebar"] [data-baseweb="select"] span,
        [data-testid="stSidebar"] [data-baseweb="input"] input,
        [data-testid="stSidebar"] .stDateInput input {
            color: #111827 !important;
            -webkit-text-fill-color: #111827 !important;
        }
        [data-testid="stSidebar"] [data-baseweb="select"] svg,
        [data-testid="stSidebar"] [data-baseweb="input"] svg {
            color: #334155 !important;
            fill: #334155 !important;
        }
        [data-testid="stSidebar"] [data-baseweb="tag"] {
            background-color: #1463ff !important;
        }
        [data-testid="stSidebar"] [data-baseweb="tag"] span {
            color: #ffffff !important;
        }
        [data-baseweb="popover"] *,
        [role="listbox"] * {
            color: #111827 !important;
        }
        [data-baseweb="popover"],
        [data-baseweb="popover"] > div,
        [role="listbox"],
        [role="option"] {
            background-color: #ffffff !important;
        }
        [role="option"],
        [role="option"] *,
        [data-baseweb="menu"] *,
        [data-baseweb="popover"] li,
        [data-baseweb="popover"] li * {
            color: #111827 !important;
            -webkit-text-fill-color: #111827 !important;
        }
        [role="option"]:hover,
        [role="option"][aria-selected="true"] {
            background-color: #dbeafe !important;
        }
        [data-baseweb="select"] div[aria-selected="true"],
        [data-baseweb="select"] div[role="option"] {
            color: #111827 !important;
            -webkit-text-fill-color: #111827 !important;
        }
        [data-testid="stSidebar"] .stButton button {
            background: #22c55e;
            color: #052e16 !important;
            border: 0;
            font-weight: 800;
        }
        h1, h2, h3 { letter-spacing: 0; }
        .hero {
            padding: 22px 26px;
            border-radius: 8px;
            background: linear-gradient(135deg, #102a6b 0%, #1463ff 52%, #00a7c8 100%);
            color: white;
            box-shadow: 0 18px 45px rgba(16, 42, 107, .22);
            margin-bottom: 18px;
        }
        .hero h1 { margin: 0; font-size: 2.1rem; line-height: 1.1; }
        .hero p { margin: 8px 0 0; color: rgba(255,255,255,.88); font-size: 1rem; }
        .metric-card {
            min-height: 132px;
            padding: 18px 18px 16px;
            border-radius: 8px;
            background: var(--card);
            border: 1px solid rgba(20, 99, 255, .10);
            box-shadow: 0 14px 30px rgba(15, 23, 42, .10);
        }
        .metric-title {
            color: var(--muted);
            text-transform: uppercase;
            font-size: .74rem;
            font-weight: 800;
            letter-spacing: .05em;
            margin-bottom: 6px;
        }
        .metric-value {
            color: var(--ink);
            font-size: 2rem;
            line-height: 1.1;
            font-weight: 850;
            white-space: nowrap;
        }
        .metric-sub {
            color: var(--muted);
            font-size: .85rem;
            margin-top: 8px;
        }
        .business-card {
            min-height: 156px;
            padding: 20px 22px;
            border-radius: 8px;
            color: white;
            box-shadow: 0 18px 36px rgba(15, 23, 42, .18);
        }
        .business-card * { color: white !important; }
        .business-cza {
            background: linear-gradient(135deg, #075985 0%, #1463ff 58%, #00a7c8 100%);
        }
        .business-ung {
            background: linear-gradient(135deg, #065f46 0%, #12b76a 60%, #84cc16 100%);
        }
        .business-title {
            font-size: .82rem;
            font-weight: 850;
            text-transform: uppercase;
            letter-spacing: .06em;
            opacity: .9;
        }
        .business-value {
            font-size: 2.5rem;
            line-height: 1.05;
            font-weight: 900;
            margin-top: 8px;
            white-space: nowrap;
        }
        .business-sub {
            margin-top: 10px;
            font-size: .95rem;
            opacity: .92;
        }
        .accent-blue { border-top: 5px solid var(--blue); }
        .accent-green { border-top: 5px solid var(--green); }
        .accent-orange { border-top: 5px solid var(--orange); }
        .accent-violet { border-top: 5px solid var(--violet); }
        .accent-red { border-top: 5px solid var(--red); }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 8px;
            box-shadow: 0 10px 24px rgba(15, 23, 42, .08);
        }
        .stTabs [data-baseweb="tab-list"] { gap: 8px; }
        .stTabs [data-baseweb="tab"] {
            border-radius: 8px;
            background: #dbeafe;
            padding: 10px 16px;
            border: 1px solid rgba(20, 99, 255, .18);
        }
        .stTabs [data-baseweb="tab"] p {
            color: #12356f !important;
            font-weight: 800;
        }
        .stTabs [aria-selected="true"] {
            background: #1463ff !important;
            border-color: #1463ff !important;
        }
        .stTabs [aria-selected="true"] p {
            color: #ffffff !important;
        }
        .exec-wrap {
            background: #ffffff;
            border: 1px solid #111827;
            border-radius: 6px;
            padding: 10px;
            margin: 10px 0 28px;
            overflow-x: auto;
            box-shadow: 0 10px 22px rgba(15, 23, 42, .08);
        }
        .exec-title {
            font-weight: 900;
            color: #111827 !important;
            margin: 4px 0 8px;
            text-transform: uppercase;
        }
        table.exec-table {
            border-collapse: collapse;
            width: 100%;
            min-width: 760px;
            font-family: Arial, sans-serif;
            font-size: 14px;
        }
        table.exec-table th {
            background: #28549a;
            color: #ffffff !important;
            border: 1px solid #111827;
            padding: 5px 7px;
            text-align: center;
            font-weight: 900;
        }
        table.exec-table td {
            border: 1px solid #111827;
            padding: 4px 7px;
            text-align: right;
            color: #111827 !important;
            background: #ffffff;
            font-weight: 700;
        }
        table.exec-table td:first-child {
            background: #2f5ea8;
            color: #ffffff !important;
            text-align: center;
            font-weight: 900;
            white-space: nowrap;
        }
        table.exec-table tr.total-row td {
            font-weight: 950;
            background: #f8fafc;
        }
        table.exec-table tr.total-row td:first-child {
            background: #214986;
            color: #ffffff !important;
        }
        .planner-note {
            background: #e0f2fe;
            padding: 8px 12px;
            border-radius: 6px;
            color: #0f172a !important;
            font-weight: 800;
            margin-bottom: 8px;
        }
        table.planner-table {
            border-collapse: collapse;
            width: 100%;
            min-width: 840px;
            font-size: 14px;
        }
        table.planner-table th {
            background: #0070c0;
            color: #ffffff !important;
            border: 1px solid #111827;
            padding: 5px 7px;
            text-align: center;
            font-weight: 900;
        }
        table.planner-table th.yellow {
            background: #ffd966;
            color: #111827 !important;
        }
        table.planner-table td {
            border: 1px solid #111827;
            padding: 4px 7px;
            text-align: right;
            color: #111827 !important;
            background: #ffffff;
            font-weight: 700;
        }
        table.planner-table td:first-child {
            background: #d9e2f3;
            color: #111827 !important;
            text-align: left;
            font-weight: 900;
            white-space: nowrap;
        }
        table.planner-table tr.total-row td:first-child,
        table.planner-table tr.mesa-row td {
            background: #2f5ea8;
            color: #ffffff !important;
            text-align: center;
            font-weight: 900;
        }
        table.planner-table td.good { background: #c6efce; color: #006100 !important; }
        table.planner-table td.bad { background: #ffc7ce; color: #9c0006 !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", str(value))
        if not unicodedata.combining(char)
    )


def clean_name(value: str) -> str:
    value = strip_accents(value).strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_")


def make_unique_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for column in columns:
        base = clean_name(column) or "columna"
        seen[base] = seen.get(base, 0) + 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def latest_file_in_folder(folder: Path) -> Path | None:
    return latest_daily_file_in_folder(folder)


def latest_matching_file(folder: Path, include_terms: tuple[str, ...], exclude_terms: tuple[str, ...] = ()) -> Path | None:
    if not folder.exists():
        return None
    include_terms = tuple(clean_name(term) for term in include_terms)
    exclude_terms = tuple(clean_name(term) for term in exclude_terms)
    files = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and path.suffix.lower() in VALID_EXTENSIONS
        and all(term in clean_name(path.stem) for term in include_terms)
        and not any(term in clean_name(path.stem) for term in exclude_terms)
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def latest_daily_file_in_folder(folder: Path) -> Path | None:
    preferred = latest_matching_file(folder, ("venta",), ("anual",))
    if preferred is not None:
        return preferred
    return latest_matching_file(folder, tuple(), ("anual",))


def latest_annual_file_in_folder(folder: Path) -> Path | None:
    return latest_matching_file(folder, ("anual",))


def latest_customer_file_in_folder(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    files = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in CLIENT_EXTENSIONS
        and any(term in clean_name(path.stem) for term in ("cliente", "clientes"))
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def latest_objectives_file_in_folder(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    files = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in CLIENT_EXTENSIONS
        and any(term in clean_name(path.stem) for term in ("objet", "sensibilizacion"))
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def latest_planner_objectives_file_in_folder(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    terms = ("objet", "planificador", "focos")
    files = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in CLIENT_EXTENSIONS
        and any(term in clean_name(path.stem) for term in terms)
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def latest_auxiliary_file_in_folder(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    files = [
        path
        for path in folder.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in CLIENT_EXTENSIONS
        and "auxiliar" in clean_name(path.stem)
    ]
    return max(files, key=lambda path: path.stat().st_mtime) if files else None


def parse_argentine_number(series: pd.Series) -> pd.Series:
    text = (
        series.astype("string")
        .str.strip()
        .str.replace("%", "", regex=False)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True)
    )
    return pd.to_numeric(text, errors="coerce")


def parse_period_date(raw: pd.Series) -> pd.Series:
    month_map = {
        "ene": "jan",
        "feb": "feb",
        "mar": "mar",
        "abr": "apr",
        "may": "may",
        "jun": "jun",
        "jul": "jul",
        "ago": "aug",
        "sep": "sep",
        "set": "sep",
        "oct": "oct",
        "nov": "nov",
        "dic": "dec",
    }

    text = raw.astype("string").str.strip().str.lower()
    text = text.str.replace(r"^\(\d+\)\s*", "", regex=True)
    for spanish, english in month_map.items():
        text = text.str.replace(fr"\b{spanish}\b", english, regex=True)
    parsed = pd.Series(pd.NaT, index=raw.index, dtype="datetime64[ns]")
    for date_format in ("%d-%b-%y", "%d-%b-%Y", "%d %b %y", "%d %b %Y", "%d/%m/%y", "%d/%m/%Y"):
        missing = parsed.isna()
        if not missing.any():
            break
        parsed.loc[missing] = pd.to_datetime(text.loc[missing], format=date_format, errors="coerce")
    if parsed.isna().any():
        parsed.loc[parsed.isna()] = pd.to_datetime(text.loc[parsed.isna()], errors="coerce", dayfirst=True)
    return parsed.dt.normalize()


def col_by_position(df: pd.DataFrame, letter: str) -> pd.Series:
    index = excel_col_to_index(letter)
    if index >= len(df.columns):
        return pd.Series(pd.NA, index=df.index, dtype="object")
    return df.iloc[:, index]


def first_present(df: pd.DataFrame, names: list[str], fallback_letter: str | None = None) -> pd.Series:
    for name in names:
        if name in df.columns:
            return df[name]
    if fallback_letter:
        return col_by_position(df, fallback_letter)
    return pd.Series(pd.NA, index=df.index, dtype="object")


def mesa_from_promoter(series: pd.Series) -> pd.Series:
    names = series.fillna("").astype(str).str.strip().str.upper()
    return names.map(PROMOTER_MESA_MAP).fillna("Sin mesa")


def load_source_from_path(
    path_text: str,
    modified_ns: int,
    normalization_version: int = NORMALIZATION_VERSION,
) -> tuple[pd.DataFrame, SourceInfo]:
    path = Path(path_text)
    raw = read_tabular(path)
    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    return normalize(raw), info


def load_annual_source_from_path(
    path_text: str,
    modified_ns: int,
    normalization_version: int = NORMALIZATION_VERSION,
) -> tuple[pd.DataFrame, SourceInfo]:
    path = Path(path_text)
    raw = read_tabular(path)
    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    return normalize(raw), info


def load_source_from_upload(
    name: str,
    content: bytes,
    normalization_version: int = NORMALIZATION_VERSION,
) -> tuple[pd.DataFrame, SourceInfo]:
    raw = read_tabular(io.BytesIO(content))
    info = SourceInfo(label=name, path=None, modified=None)
    return normalize(raw), info


def read_tabular(source: str | Path | io.BytesIO) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            if hasattr(source, "seek"):
                source.seek(0)
            return pd.read_csv(
                source,
                sep="\t",
                dtype="string",
                encoding=encoding,
                engine="c",
            )
        except (UnicodeDecodeError, pd.errors.ParserError) as exc:
            last_error = exc
    raise RuntimeError(f"No pude detectar la codificacion del archivo: {last_error}")


def classify_customer_channel(value: str | None) -> str:
    text = strip_accents("" if value is None else str(value)).upper()
    if "AUTOSERV" in text:
        return "AUTOSERVICIO"
    if "MAYOR" in text:
        return "MAYORISTA"
    if "REFRIG" in text:
        return "REF"
    if any(term in text for term in ("TRADICIONAL", "KIOSCO", "CADENITA", "LISTA UNICA")):
        return "K+T"
    return "NO"


def read_excel_selected_columns(path: Path, sheet_name: str, wanted_columns: list[str], header_row: int = 2) -> pd.DataFrame:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return pd.DataFrame(columns=wanted_columns)
    return read_excel_selected_columns_from_sheet(workbook[sheet_name], wanted_columns, header_row)


def read_excel_selected_columns_from_sheet(sheet, wanted_columns: list[str], header_row: int = 2) -> pd.DataFrame:
    header_values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
    header_map = {str(value).strip(): index for index, value in enumerate(header_values) if value is not None}
    indexes = [header_map[column] for column in wanted_columns if column in header_map]
    names = [column for column in wanted_columns if column in header_map]
    if not indexes:
        return pd.DataFrame(columns=wanted_columns)

    rows = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        rows.append([values[index] if index < len(values) else None for index in indexes])
    return pd.DataFrame(rows, columns=names, dtype="string")


def workbook_sheet_name(path: Path, contains: str) -> str | None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    target = clean_name(contains)
    return next((name for name in workbook.sheetnames if target in clean_name(name)), None)


def workbook_sheet_name_from_workbook(workbook, contains: str) -> str | None:
    target = clean_name(contains)
    return next((name for name in workbook.sheetnames if target in clean_name(name)), None)


def load_customer_channels(
    path_text: str,
    modified_ns: int,
    customer_channel_version: int = CUSTOMER_CHANNEL_VERSION,
) -> tuple[pd.DataFrame, SourceInfo]:
    path = Path(path_text)
    if path.suffix.lower() in {".xlsx", ".xls"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        customers = (
            read_excel_selected_columns_from_sheet(workbook["Clientes"], ["Cliente", "Lista de precios", "Subcanal MKT"])
            if "Clientes" in workbook.sheetnames
            else pd.DataFrame(columns=["Cliente", "Lista de precios", "Subcanal MKT"])
        )
        price_lists = (
            read_excel_selected_columns_from_sheet(workbook["Listas de precios"], ["Código", "Descripción"])
            if "Listas de precios" in workbook.sheetnames
            else pd.DataFrame(columns=["Código", "Descripción"])
        )
        hierarchy_sheet = workbook_sheet_name_from_workbook(workbook, "Jerarqu")
        hierarchy = (
            read_excel_selected_columns_from_sheet(workbook[hierarchy_sheet], ["Código", "Segmento MKT", "Canal MKT", "Subcanal MKT"])
            if hierarchy_sheet
            else pd.DataFrame()
        )
    else:
        customers = read_tabular(path)
        price_lists = pd.DataFrame(columns=["Código", "Descripción"])
        hierarchy = pd.DataFrame()

    customers = customers[customers.get("Cliente").notna()].copy()
    customers = customers[customers["Cliente"].astype(str).str.upper() != "ENTERO"]
    customers["cliente_codigo"] = pd.to_numeric(customers["Cliente"], errors="coerce").astype("Int64").astype("string")
    customers["lista_precio_codigo"] = pd.to_numeric(customers.get("Lista de precios"), errors="coerce").astype("Int64").astype("string")
    customers["subcanal_mkt_codigo"] = pd.to_numeric(customers.get("Subcanal MKT"), errors="coerce").astype("Int64").astype("string")

    if not price_lists.empty and {"Código", "Descripción"}.issubset(price_lists.columns):
        price_lists = price_lists[price_lists["Código"].notna()].copy()
        price_lists["lista_precio_codigo"] = pd.to_numeric(price_lists["Código"], errors="coerce").astype("Int64").astype("string")
        price_lists = price_lists[["lista_precio_codigo", "Descripción"]].rename(
            columns={"Descripción": "lista_precio_descripcion"}
        )
        customers = customers.merge(price_lists, on="lista_precio_codigo", how="left")
    else:
        customers["lista_precio_descripcion"] = customers["lista_precio_codigo"]

    if not hierarchy.empty and {"Código", "Segmento MKT", "Canal MKT", "Subcanal MKT"}.issubset(hierarchy.columns):
        hierarchy = hierarchy[hierarchy["Código"].notna()].copy()
        hierarchy["subcanal_mkt_codigo"] = pd.to_numeric(hierarchy["Código"], errors="coerce").astype("Int64").astype("string")
        hierarchy = hierarchy[
            ["subcanal_mkt_codigo", "Segmento MKT", "Canal MKT", "Subcanal MKT"]
        ].rename(
            columns={
                "Segmento MKT": "segmento_mkt",
                "Canal MKT": "canal_mkt",
                "Subcanal MKT": "subcanal_mkt",
            }
        )
        customers = customers.merge(hierarchy, on="subcanal_mkt_codigo", how="left")
    else:
        customers["segmento_mkt"] = ""
        customers["canal_mkt"] = ""
        customers["subcanal_mkt"] = ""

    channel_text = (
        customers["lista_precio_descripcion"].fillna("")
        + " "
        + customers["segmento_mkt"].fillna("")
        + " "
        + customers["canal_mkt"].fillna("")
        + " "
        + customers["subcanal_mkt"].fillna("")
    )
    customers["canal_maestro"] = channel_text.apply(classify_customer_channel)
    result = customers[
        ["cliente_codigo", "canal_maestro", "lista_precio_descripcion", "segmento_mkt", "canal_mkt", "subcanal_mkt"]
    ].drop_duplicates("cliente_codigo")
    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    return result, info


def apply_customer_channels(df: pd.DataFrame, customer_channels: pd.DataFrame | None) -> pd.DataFrame:
    if customer_channels is None or customer_channels.empty or "cliente_codigo" not in df.columns:
        return df
    result = df.copy()
    result["cliente_codigo"] = pd.to_numeric(result["cliente_codigo"], errors="coerce").astype("Int64").astype("string")
    result = result.merge(customer_channels, on="cliente_codigo", how="left")
    result["canal"] = result["canal_maestro"].fillna("NO")
    result["lista_precio_descripcion"] = result["lista_precio_descripcion"].fillna("Sin lista")
    for column in ["segmento_mkt", "canal_mkt", "subcanal_mkt"]:
        if column in result.columns:
            result[column] = result[column].fillna("Sin dato")
    return result.drop(columns=["canal_maestro"])


def key_text(value: str | None) -> str:
    return strip_accents("" if value is None or pd.isna(value) else str(value)).upper().strip()


def normalize_beer_segment(value: str | None) -> str:
    text = key_text(value)
    if text == "CORE PLUS":
        return "CVZA CORE +"
    if text == "VALUE":
        return "CVZA VALUE"
    if text == "HE":
        return "CVZA HE"
    if text == "CORE":
        return "CVZA CORE"
    return "CVZA SIN SEGMENTO"


def load_auxiliary_segments(
    path_text: str,
    modified_ns: int,
    segment_version: int = SEGMENT_VERSION,
) -> tuple[dict[str, pd.DataFrame], SourceInfo]:
    path = Path(path_text)
    aux = pd.read_excel(path, sheet_name="PIVOT", dtype="string")

    beer = aux[["MARCA", "SEGMENTO"]].dropna(subset=["MARCA"]).copy()
    beer["marca_key"] = beer["MARCA"].apply(key_text)
    beer["segmento_cerveza"] = beer["SEGMENTO"].apply(normalize_beer_segment)
    beer = beer[["marca_key", "segmento_cerveza"]].drop_duplicates("marca_key")

    ung = aux[["Marca", "Calibre", "UNG TOP"]].dropna(subset=["Marca", "Calibre"]).copy()
    ung["marca_key"] = ung["Marca"].apply(key_text)
    ung["calibre_key"] = ung["Calibre"].apply(key_text)
    ung["segmento_ung"] = np.where(
        ung["UNG TOP"].fillna("").astype(str).str.upper().str.contains("UNG TOP"),
        "UNG TOP",
        "UNG SIN TOP",
    )
    ung = ung[["marca_key", "calibre_key", "segmento_ung"]].drop_duplicates(["marca_key", "calibre_key"])

    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    return {"beer": beer, "ung": ung}, info


def apply_auxiliary_segments(df: pd.DataFrame, aux_segments: dict[str, pd.DataFrame] | None) -> pd.DataFrame:
    result = df.copy()
    result["division_informe"] = result.get("unidad_negocio", pd.Series("Sin negocio", index=result.index)).astype(str)
    if aux_segments is None:
        result.loc[result["unidad_negocio"] == "CZA", "division_informe"] = "CVZA SIN SEGMENTO"
        result.loc[result["unidad_negocio"] == "UNG", "division_informe"] = "UNG SIN TOP"
        return result

    result["marca_key"] = result["marca"].apply(key_text) if "marca" in result.columns else ""
    result["calibre_key"] = result["calibre"].apply(key_text) if "calibre" in result.columns else ""

    beer = aux_segments.get("beer", pd.DataFrame())
    if not beer.empty:
        result = result.merge(beer, on="marca_key", how="left")
        cza_mask = result["unidad_negocio"] == "CZA"
        result.loc[cza_mask, "division_informe"] = result.loc[cza_mask, "segmento_cerveza"].fillna("CVZA SIN SEGMENTO")

    ung = aux_segments.get("ung", pd.DataFrame())
    if not ung.empty:
        result = result.merge(ung, on=["marca_key", "calibre_key"], how="left")
        ung_mask = result["unidad_negocio"] == "UNG"
        result.loc[ung_mask, "division_informe"] = result.loc[ung_mask, "segmento_ung"].fillna("UNG SIN TOP")

    result.loc[~result["unidad_negocio"].isin(["CZA", "UNG"]), "division_informe"] = result.loc[
        ~result["unidad_negocio"].isin(["CZA", "UNG"]), "unidad_negocio"
    ]
    return result.drop(columns=[col for col in ["marca_key", "calibre_key", "segmento_cerveza", "segmento_ung"] if col in result.columns])


def load_objectives(path_text: str, modified_ns: int) -> tuple[pd.DataFrame, SourceInfo]:
    path = Path(path_text)
    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    if path.suffix.lower() in {".xlsx", ".xls"}:
        matrix_result = parse_report_objectives_matrix(path)
        if not matrix_result.empty:
            return matrix_result, info
        objectives = pd.read_excel(path, dtype="string")
    else:
        objectives = read_tabular(path)
    objectives = objectives.copy()
    objectives.columns = make_unique_columns(list(objectives.columns))

    section_col = next((col for col in objectives.columns if col in {"seccion", "tabla", "grupo"}), None)
    item_col = next((col for col in objectives.columns if col in {"item", "division", "mesa", "calibre", "canal"}), None)
    objective_col = next((col for col in objectives.columns if "objet" in col), None)
    if item_col is None and len(objectives.columns) >= 1:
        item_col = objectives.columns[0]
    if objective_col is None and len(objectives.columns) >= 2:
        objective_col = objectives.columns[1]
    if item_col is None or objective_col is None:
        raise ValueError("El archivo de objetivos debe tener al menos columnas item y objetivo.")

    result = pd.DataFrame(
        {
            "seccion": objectives[section_col].fillna("").astype(str).str.strip().str.upper()
            if section_col
            else "",
            "item": objectives[item_col].fillna("").astype(str).str.strip().str.upper(),
            "OBJ VTAS": parse_argentine_number(objectives[objective_col]),
        }
    ).dropna(subset=["OBJ VTAS"])

    return result, info


def parse_report_objectives_matrix(path: Path) -> pd.DataFrame:
    source = pd.read_excel(path, sheet_name=0, header=None)
    header_index = None
    for index, row in source.iterrows():
        values = row.fillna("").astype(str).tolist()
        if any("descripcion" in clean_name(value) for value in values) and sum("-" in value for value in values) >= 2:
            header_index = index
            break
    if header_index is None:
        return pd.DataFrame(columns=["seccion", "item", "OBJ VTAS"])

    header = source.loc[header_index].fillna("").astype(str)
    total_col = next((col for col, value in header.items() if clean_name(value) == "total"), None)
    subtotal_col = next((col for col, value in header.items() if "subtotal" in clean_name(value)), None)
    vendor_columns = [
        col
        for col, value in header.items()
        if col >= 2 and "-" in str(value) and normalize_vendor_name(value)
    ]

    rows_by_code: dict[str, float] = {}
    for _, row in source.loc[header_index + 1 :].iterrows():
        code_match = re.search(r"(\d+)", str(row.iloc[0] if len(row) else ""))
        if not code_match:
            continue
        code = code_match.group(1)
        if total_col is not None:
            value = parse_objective_cell(row.get(total_col))
        elif subtotal_col is not None:
            value = parse_objective_cell(row.get(subtotal_col))
        else:
            values = [parse_objective_cell(row.get(col)) for col in vendor_columns]
            value = float(np.nansum(values)) if values else np.nan
        if not pd.isna(value):
            rows_by_code[code] = float(value)

    if not rows_by_code:
        return pd.DataFrame(columns=["seccion", "item", "OBJ VTAS"])

    total_cvza = rows_by_code.get("2218", np.nan)
    total_ung = rows_by_code.get("19341", np.nan)
    aguas = rows_by_code.get("18743", np.nan)
    premium_codes = [code for code in rows_by_code if code not in {"2218", "19341", "18743"}]
    premium = float(np.nansum([rows_by_code[code] for code in premium_codes])) if premium_codes else np.nan
    core_value = total_cvza - premium if not pd.isna(total_cvza) and not pd.isna(premium) else np.nan

    rows = [
        {"seccion": "", "item": "TOTAL CVZA", "OBJ VTAS": total_cvza},
        {"seccion": "", "item": "TOTAL UNG", "OBJ VTAS": total_ung},
        {"seccion": "", "item": "CZA", "OBJ VTAS": total_cvza},
        {"seccion": "", "item": "UNG", "OBJ VTAS": total_ung},
        {"seccion": "", "item": "AGUAS ECO", "OBJ VTAS": aguas},
        {"seccion": "", "item": "CVZA HE", "OBJ VTAS": premium},
        {"seccion": "", "item": "CVZA CORE + VALUE", "OBJ VTAS": core_value},
    ]

    return pd.DataFrame(rows).dropna(subset=["OBJ VTAS"])


def normalize_vendor_name(value: object) -> str:
    text = str(value or "").strip()
    if "-" in text:
        text = text.split("-", 1)[1]
    normalized = re.sub(r"\s+", " ", text).strip().upper()
    return VENDOR_NAME_ALIASES.get(normalized, normalized)


def normalize_planner_focus(value: object) -> str:
    cleaned = clean_name(value).replace("_", " ").upper().strip()
    for alias, focus_name in PLANNER_OBJECTIVE_ALIASES.items():
        if alias in cleaned:
            return focus_name
    for focus_name in PLANNER_FOCUS_RULES:
        if clean_name(focus_name).replace("_", " ").upper() in cleaned:
            return focus_name
    return str(value or "").strip()


def parse_objective_cell(value: object) -> float:
    if value is None or pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip()
    if not text:
        return np.nan
    if "," in text:
        return float(parse_argentine_number(pd.Series([text])).iloc[0])
    return float(pd.to_numeric(text, errors="coerce"))


def parse_cross_planner_objectives(path: Path) -> pd.DataFrame:
    source = pd.read_excel(path, sheet_name=0, header=None)
    header_index = None
    for index, row in source.iterrows():
        values = row.fillna("").astype(str).tolist()
        if any("descripcion" in clean_name(value) for value in values) and sum("-" in value for value in values) >= 2:
            header_index = index
            break
    if header_index is None:
        raise ValueError("No pude detectar la fila de vendedores en objetivos.")
    header = source.loc[header_index]
    vendor_columns = {
        col: normalize_vendor_name(value)
        for col, value in header.items()
        if col >= 2 and "-" in str(value) and normalize_vendor_name(value)
    }
    rows: list[dict[str, object]] = []
    for _, row in source.loc[header_index + 1 :].iterrows():
        focus_name = normalize_planner_focus(row.iloc[1] if len(row) > 1 else "")
        if focus_name not in PLANNER_FOCUS_RULES:
            continue
        for col, vendor in vendor_columns.items():
            value = parse_objective_cell(row.get(col))
            if not pd.isna(value):
                rows.append({"promotor": vendor, "foco": focus_name, "objetivo_mes": float(value)})
    return pd.DataFrame(rows)


def google_sheet_export_url(raw_url: str) -> str:
    raw_url = str(raw_url or "").strip()
    if not raw_url:
        return ""
    if raw_url.lower().endswith((".xlsx", ".xls")) or "output=xlsx" in raw_url.lower():
        return raw_url
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", raw_url)
    if match:
        return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
    return raw_url


def planner_sheet_url() -> str:
    url = secret_or_env("PLANNER_GOOGLE_SHEET_URL")
    return "" if url.upper().startswith("PEGAR_") else url


def planner_webapp_url() -> str:
    url = secret_or_env("PLANNER_WEBAPP_URL")
    return "" if url.upper().startswith("PEGAR_") else url


def spreadsheet_col_letter(col_number: int) -> str:
    letters = ""
    while col_number > 0:
        col_number, remainder = divmod(col_number - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def spreadsheet_cell(row_index: int, col_index: int) -> str:
    return f"{spreadsheet_col_letter(col_index + 1)}{row_index + 1}"


def parse_planning_sheet_workbook(workbook: dict[str, pd.DataFrame], selected_date: pd.Timestamp) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for sheet_name, sheet in workbook.items():
        raw = sheet.fillna("")
        for row_idx in range(len(raw)):
            for col_idx in range(len(raw.columns)):
                focus_name = normalize_planner_focus(raw.iat[row_idx, col_idx])
                if focus_name not in PLANNER_FOCUS_RULES:
                    continue
                header_idx = None
                for candidate in range(row_idx, min(row_idx + 8, len(raw))):
                    row_values = [clean_name(value) for value in raw.iloc[candidate].tolist()]
                    has_promoter = any("promotor" in value for value in row_values)
                    has_plan = any("planific" in value for value in row_values)
                    if has_promoter and has_plan:
                        header_idx = candidate
                        break
                if header_idx is None:
                    continue
                header = [clean_name(value) for value in raw.iloc[header_idx].tolist()]
                search_from = max(0, col_idx - 1)
                search_to = min(len(header), col_idx + 5)
                header_scope = list(enumerate(header[search_from:search_to], start=search_from))
                promoter_col = next((i for i, value in header_scope if "promotor" in value), None)
                plan_col = next((i for i, value in header_scope if "planific" in value), None)
                if promoter_col is None or plan_col is None:
                    continue
                objective_col = next((i for i, value in header_scope if "objet" in value), None)
                for detail_idx in range(header_idx + 1, len(raw)):
                    promoter = normalize_vendor_name(raw.iat[detail_idx, promoter_col])
                    if not promoter:
                        continue
                    promoter_clean = clean_name(promoter)
                    if "total" in promoter_clean or "promotor" in promoter_clean:
                        break
                    plan_value = parse_objective_cell(raw.iat[detail_idx, plan_col])
                    objective_value = parse_objective_cell(raw.iat[detail_idx, objective_col]) if objective_col is not None else np.nan
                    rows.append(
                        {
                            "fecha": pd.Timestamp(selected_date).normalize(),
                            "foco": focus_name,
                            "promotor": promoter,
                            "planificado": plan_value,
                            "objetivo_mes": objective_value,
                            "supervisor": str(sheet_name),
                            "hoja_origen": str(sheet_name),
                            "celda_planificacion": spreadsheet_cell(detail_idx, plan_col),
                            "celda_objetivo": spreadsheet_cell(detail_idx, objective_col) if objective_col is not None else "",
                        }
                    )
    if not rows:
        return pd.DataFrame(columns=PLANNER_COLUMNS + ["objetivo_mes", "supervisor", "hoja_origen", "celda_planificacion", "celda_objetivo"])
    result = pd.DataFrame(rows)
    result = result.drop_duplicates(["fecha", "foco", "promotor"], keep="last")
    return result


@st.cache_data(show_spinner=False, ttl=300)
def load_remote_planning_sheet(raw_url: str, selected_date_text: str) -> pd.DataFrame:
    url = google_sheet_export_url(raw_url)
    if not url:
        return pd.DataFrame(columns=PLANNER_COLUMNS + ["objetivo_mes", "supervisor"])
    workbook = pd.read_excel(url, sheet_name=None, header=None)
    return parse_planning_sheet_workbook(workbook, pd.Timestamp(selected_date_text))


def load_planner_objectives(path_text: str, modified_ns: int) -> tuple[pd.DataFrame, SourceInfo]:
    path = Path(path_text)
    if path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            result = parse_cross_planner_objectives(path)
        except Exception:
            source = pd.read_excel(path, dtype="string")
            source.columns = make_unique_columns(list(source.columns))
            vendor_col = next((col for col in source.columns if "vendedor" in col or "promotor" in col), None)
            focus_col = next((col for col in source.columns if "foco" in col or "segmento" in col or "division" in col), None)
            obj_col = next((col for col in source.columns if "objet" in col or "plan" in col), None)
            if vendor_col is None or obj_col is None:
                raise
            result = pd.DataFrame(
                {
                    "promotor": source[vendor_col].map(normalize_vendor_name),
                    "foco": source[focus_col].map(normalize_planner_focus) if focus_col else "",
                    "objetivo_mes": parse_argentine_number(source[obj_col]),
                }
            )
    else:
        source = read_tabular(path)
        source.columns = make_unique_columns(list(source.columns))
        vendor_col = next((col for col in source.columns if "vendedor" in col or "promotor" in col), None)
        focus_col = next((col for col in source.columns if "foco" in col or "segmento" in col or "division" in col), None)
        obj_col = next((col for col in source.columns if "objet" in col or "plan" in col), None)
        if vendor_col is None or obj_col is None:
            raise ValueError("El archivo de objetivos por vendedor debe tener vendedor/promotor y objetivo.")
        result = pd.DataFrame(
            {
                "promotor": source[vendor_col].map(normalize_vendor_name),
                "foco": source[focus_col].map(normalize_planner_focus) if focus_col else "",
                "objetivo_mes": parse_argentine_number(source[obj_col]),
            }
        )
    result = result.dropna(subset=["objetivo_mes"])
    result = result[result["promotor"].astype(str).str.strip() != ""].copy()
    info = SourceInfo(
        label=path.name,
        path=str(path),
        modified=pd.to_datetime(modified_ns, unit="ns").strftime("%d/%m/%Y %H:%M"),
    )
    return result, info


def normalize(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()
    df.columns = make_unique_columns(list(df.columns))

    fecha = parse_period_date(first_present(df, ["descripcion_periodo"], "C"))
    if fecha.isna().all():
        fecha = parse_period_date(first_present(df, ["periodos"], "A"))

    normalized = pd.DataFrame(
        {
            "fecha": fecha,
            "periodo_codigo": first_present(df, ["cod_periodo"], "B"),
            "cliente_codigo": first_present(df, ["cod_cliente"], "E"),
            "cliente": first_present(df, ["descripcion"], "F"),
            "ruta_codigo": first_present(df, ["ruta"], "I"),
            "ruta": first_present(df, ["descripcion_1"], "J"),
            "vendedor_codigo": first_present(df, ["vendedor"], "O"),
            "vendedor": first_present(df, ["descripcion_vendedor"], "P"),
            "marca": first_present(df, ["descripcion_3"], "U"),
            "calibre": col_by_position(df, "X"),
            "division_codigo": first_present(df, ["division"], "Z"),
            "division": first_present(df, ["descripcion_5"], "AA"),
            "canal_codigo": first_present(df, ["ramo"], "M"),
            "canal": first_present(df, ["descripcion_ramo"], "N"),
            "negocio_codigo": first_present(df, ["unidad_de_negocio"], "AI"),
            "negocio": first_present(df, ["descripcion_8"], "AJ"),
            "hl": parse_argentine_number(col_by_position(df, "AO")),
        }
    )

    normalized["ruta"] = (
        normalized["ruta_codigo"].fillna("").astype(str).str.strip()
        + " - "
        + normalized["ruta"].fillna("").astype(str).str.strip()
    ).str.strip(" -")
    normalized["vendedor"] = normalized["vendedor"].fillna("Sin vendedor").astype(str).str.strip()
    normalized["supervisor"] = normalized["vendedor"]
    normalized["promotor"] = normalized["vendedor"]
    normalized["calibre"] = normalized["calibre"].fillna("Sin calibre").astype(str).str.strip()
    normalized["division"] = normalized["division"].fillna("Sin division").astype(str).str.strip()
    normalized["canal"] = normalized["canal"].fillna("Sin canal").astype(str).str.strip()
    normalized["negocio"] = normalized["negocio"].fillna("Sin negocio").astype(str).str.strip()
    normalized["mesa"] = mesa_from_promoter(normalized["promotor"])
    normalized["unidad_negocio"] = np.select(
        [
            normalized["negocio"].str.contains("UNG", case=False, na=False),
            normalized["negocio"].str.contains("CZA|CERVEZ", case=False, na=False),
        ],
        ["UNG", "CZA"],
        default=normalized["negocio"].fillna("Otro").astype(str).str.strip(),
    )

    normalized = normalized.dropna(subset=["fecha"])
    normalized["hl"] = normalized["hl"].fillna(0.0)
    normalized["dia_semana"] = normalized["fecha"].dt.day_name(locale=None)
    normalized["es_habil"] = normalized["fecha"].dt.weekday <= 5
    normalized = normalized[normalized["es_habil"]].copy()
    normalized["mes"] = normalized["fecha"].dt.to_period("M").astype(str)
    normalized["dia_habil_mes"] = (
        normalized[["fecha"]]
        .drop_duplicates()
        .sort_values("fecha")
        .assign(dia_habil_mes=lambda x: x.groupby(x["fecha"].dt.to_period("M")).cumcount() + 1)
        .set_index("fecha")["dia_habil_mes"]
        .reindex(normalized["fecha"])
        .to_numpy()
    )
    return normalized


def filter_options(df: pd.DataFrame, column: str, label: str) -> list[str]:
    if column not in df.columns:
        return []
    values = sorted(df[column].dropna().astype(str).unique().tolist())
    selected = st.sidebar.multiselect(label, values, default=[])
    return selected


def ensure_analysis_columns(df: pd.DataFrame) -> pd.DataFrame:
    defaults = {
        "division": "Sin division",
        "division_informe": "Sin division",
        "unidad_negocio": "Sin negocio",
        "calibre": "Sin calibre",
        "mesa": "Sin mesa",
        "canal": "Sin canal",
        "supervisor": "Sin supervisor",
        "promotor": "Sin promotor",
    }
    result = df.copy()
    for column, default in defaults.items():
        if column not in result.columns:
            result[column] = default
        else:
            result[column] = result[column].fillna(default).astype(str).str.strip()
    return result


def apply_filters(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.Timestamp, dict[str, list[str]]]:
    df = ensure_analysis_columns(df)
    min_date = df["fecha"].min().date()
    max_data_date = df["fecha"].max()
    max_date = max_data_date.date()

    selected_date = st.sidebar.date_input(
        "Fecha de analisis",
        value=max_date,
        min_value=min_date,
        max_value=max_date,
    )
    date_range = st.sidebar.date_input(
        "Rango historico",
        value=(min_date, max_date),
        min_value=min_date,
        max_value=max_date,
    )

    filtered = df.copy()
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start, end = pd.Timestamp(date_range[0]), pd.Timestamp(date_range[1])
        filtered = filtered[(filtered["fecha"] >= start) & (filtered["fecha"] <= end)]

    dimension_filters: dict[str, list[str]] = {}
    for column, label in [
        ("division", "Division"),
        ("supervisor", "Supervisor"),
        ("promotor", "Promotor"),
        ("ruta", "Ruta"),
        ("mesa", "Mesa"),
        ("canal", "Canal"),
        ("unidad_negocio", "Negocio"),
        ("calibre", "Calibre"),
    ]:
        selected = filter_options(filtered, column, label)
        dimension_filters[column] = selected
        if selected:
            filtered = filtered[filtered[column].astype(str).isin(selected)]

    return filtered, pd.Timestamp(selected_date), dimension_filters


def apply_dimension_filters(df: pd.DataFrame, dimension_filters: dict[str, list[str]]) -> pd.DataFrame:
    filtered = df.copy()
    for column, selected in dimension_filters.items():
        if selected and column in filtered.columns:
            filtered = filtered[filtered[column].astype(str).isin(selected)]
    return filtered


def combine_current_with_history(current_df: pd.DataFrame, annual_df: pd.DataFrame | None) -> pd.DataFrame:
    if annual_df is None or annual_df.empty:
        return current_df.copy()
    if current_df.empty:
        return annual_df.copy()
    current_start = current_df["fecha"].min()
    history = annual_df[annual_df["fecha"] < current_start].copy()
    return pd.concat([history, current_df.copy()], ignore_index=True)


def window_stats(daily: pd.DataFrame, selected_date: pd.Timestamp) -> dict[int, dict[str, float]]:
    history = daily[daily["fecha"] <= selected_date].sort_values("fecha")
    result: dict[int, dict[str, float]] = {}
    for window in WINDOWS:
        values = history.tail(window)["hl"]
        result[window] = {
            "promedio": values.mean(),
            "mediana": values.median(),
            "min": values.min(),
            "max": values.max(),
            "p25": values.quantile(0.25),
            "p75": values.quantile(0.75),
        }
    return result


def average_last_business_days(daily: pd.DataFrame, selected_date: pd.Timestamp, days: int) -> float:
    history = daily[daily["fecha"] <= selected_date].sort_values("fecha")
    if history.empty:
        return np.nan
    return float(history.tail(days)["hl"].mean())


def exact_previous_month_value(
    daily: pd.DataFrame,
    selected_date: pd.Timestamp,
    months_back: int,
) -> tuple[pd.Timestamp, float]:
    target_date = selected_date - pd.DateOffset(months=months_back)
    value = daily.loc[daily["fecha"] == target_date.normalize(), "hl"]
    if value.empty:
        return target_date.normalize(), np.nan
    return target_date.normalize(), float(value.iloc[0])


def exact_previous_year_value(
    daily_aa: pd.DataFrame,
    selected_date: pd.Timestamp,
) -> tuple[pd.Timestamp, float]:
    target_date = selected_date - pd.DateOffset(years=1)
    value = daily_aa.loc[daily_aa["fecha"] == target_date.normalize(), "hl"]
    if value.empty:
        return target_date.normalize(), np.nan
    return target_date.normalize(), float(value.iloc[0])


def argentina_holidays_for_years(years: list[int]) -> set[pd.Timestamp]:
    if holidays is None:
        return set()
    holiday_dates = holidays.country_holidays("AR", years=years)
    return {pd.Timestamp(day).normalize() for day in holiday_dates.keys()}


def selling_day_weight(date_value: pd.Timestamp, holiday_dates: set[pd.Timestamp]) -> float:
    date_value = pd.Timestamp(date_value).normalize()
    if date_value in holiday_dates or date_value.weekday() == 6:
        return 0.0
    if date_value.weekday() == 5:
        return 0.5
    return 1.0


def weighted_selling_days(start: pd.Timestamp, end: pd.Timestamp) -> float:
    if pd.isna(start) or pd.isna(end) or end < start:
        return 0.0
    dates = pd.date_range(start.normalize(), end.normalize(), freq="D")
    holiday_dates = argentina_holidays_for_years(sorted(set(dates.year.tolist())))
    return float(sum(selling_day_weight(date, holiday_dates) for date in dates))


def selling_days_in_month(date_value: pd.Timestamp) -> float:
    date_value = pd.Timestamp(date_value).normalize()
    month_start = date_value.replace(day=1)
    month_end = month_start + pd.offsets.MonthEnd(0)
    return weighted_selling_days(month_start, month_end)


def selling_days_remaining_from(date_value: pd.Timestamp) -> float:
    date_value = pd.Timestamp(date_value).normalize()
    month_end = date_value.replace(day=1) + pd.offsets.MonthEnd(0)
    return weighted_selling_days(date_value, month_end)


def next_selling_day(date_value: pd.Timestamp) -> pd.Timestamp:
    candidate = pd.Timestamp(date_value).normalize() + pd.Timedelta(days=1)
    while True:
        holiday_dates = argentina_holidays_for_years([candidate.year])
        if selling_day_weight(candidate, holiday_dates) > 0:
            return candidate
        candidate += pd.Timedelta(days=1)


def projected_month_trend(accumulated_value: float, selected_date: pd.Timestamp) -> float:
    month_start = selected_date.replace(day=1)
    month_end = selected_date + pd.offsets.MonthEnd(0)
    elapsed_weight = weighted_selling_days(month_start, selected_date)
    month_weight = weighted_selling_days(month_start, month_end)
    if elapsed_weight <= 0:
        return np.nan
    return float(accumulated_value / elapsed_weight * month_weight)


def accumulated_vs_previous_year(
    current_daily: pd.DataFrame,
    annual_daily: pd.DataFrame,
    selected_date: pd.Timestamp,
) -> dict[str, float | pd.Timestamp]:
    current_start = selected_date.replace(day=1)
    previous_start = current_start - pd.DateOffset(years=1)
    previous_end = selected_date - pd.DateOffset(years=1)

    current_value = current_daily.loc[
        (current_daily["fecha"] >= current_start) & (current_daily["fecha"] <= selected_date),
        "hl",
    ].sum()
    previous_value = annual_daily.loc[
        (annual_daily["fecha"] >= previous_start) & (annual_daily["fecha"] <= previous_end),
        "hl",
    ].sum()
    projected_value = projected_month_trend(float(current_value), selected_date)
    trend = (projected_value / previous_value * 100) if previous_value else np.nan
    return {
        "current_start": current_start,
        "previous_start": previous_start,
        "previous_end": previous_end,
        "current_value": float(current_value),
        "previous_value": float(previous_value) if previous_value else np.nan,
        "projected_value": projected_value,
        "trend": trend,
    }


def previous_month_same_business_day(daily: pd.DataFrame, selected_date: pd.Timestamp) -> float:
    return previous_same_business_day_by_months(daily, selected_date, 1)


def previous_same_business_day_by_months(
    daily: pd.DataFrame,
    selected_date: pd.Timestamp,
    months_back: int,
) -> float:
    calendar = daily[["fecha"]].drop_duplicates().sort_values("fecha").copy()
    if calendar.empty:
        return np.nan
    calendar["mes_period"] = calendar["fecha"].dt.to_period("M")
    calendar["dia_habil_mes"] = calendar.groupby("mes_period").cumcount() + 1

    current = calendar[calendar["fecha"] == selected_date]
    if current.empty:
        return np.nan

    previous_month = selected_date.to_period("M") - months_back
    business_index = int(current["dia_habil_mes"].iloc[0])
    match = calendar[
        (calendar["mes_period"] == previous_month)
        & (calendar["dia_habil_mes"] == business_index)
    ]
    if match.empty:
        return np.nan
    previous_date = match["fecha"].iloc[0]
    value = daily.loc[daily["fecha"] == previous_date, "hl"]
    return float(value.iloc[0]) if not value.empty else np.nan


def format_hl(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def format_pct(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{value:+.1f}%".replace(".", ",")


def format_pct_plain(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{value:.0f}%".replace(".", ",")


def valid_nonzero(value: object) -> bool:
    if value is None or pd.isna(value):
        return False
    try:
        return float(value) != 0
    except (TypeError, ValueError):
        return False


def planner_store_path() -> Path:
    return PLANNER_DATA_DIR / PLANNER_STORE_FILE_NAME


def normalize_saved_planner(saved: pd.DataFrame) -> pd.DataFrame:
    for column in PLANNER_COLUMNS:
        if column not in saved.columns:
            saved[column] = np.nan
    saved = saved[PLANNER_COLUMNS].copy()
    saved["fecha"] = pd.to_datetime(saved["fecha"], errors="coerce")
    saved["promotor"] = saved["promotor"].map(normalize_vendor_name)
    saved["planificado"] = saved["planificado"].map(parse_objective_cell)
    return saved.dropna(subset=["fecha"])


def local_planner_load() -> pd.DataFrame:
    path = planner_store_path()
    if not path.exists():
        return pd.DataFrame(columns=PLANNER_COLUMNS)
    try:
        saved = pd.read_csv(path, sep=";", dtype="string")
    except Exception:
        return pd.DataFrame(columns=PLANNER_COLUMNS)
    return normalize_saved_planner(saved)


def local_planner_save(saved: pd.DataFrame) -> str:
    PLANNER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    output = saved.copy()
    output["fecha"] = pd.to_datetime(output["fecha"]).dt.strftime("%Y-%m-%d")
    path = planner_store_path()
    output.to_csv(path, sep=";", index=False)
    return str(path)


def load_saved_planner() -> pd.DataFrame:
    return local_planner_load()


def save_plan_to_google_sheet(selected_date: pd.Timestamp, focus_name: str, plan_df: pd.DataFrame) -> str | None:
    url = planner_webapp_url()
    if not url:
        return None
    rows = []
    for _, row in plan_df.iterrows():
        planificado = pd.to_numeric(row.get("PLANIFICADO"), errors="coerce")
        if pd.isna(planificado):
            planificado = ""
        rows.append(
            {
                "fecha": pd.Timestamp(selected_date).strftime("%Y-%m-%d"),
                "foco": focus_name,
                "promotor": normalize_vendor_name(row.get("promotor")),
                "planificado": "" if planificado == "" else float(planificado),
            }
        )
    payload = {"fecha": pd.Timestamp(selected_date).strftime("%Y-%m-%d"), "foco": focus_name, "rows": rows}
    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            body = response.read().decode("utf-8", errors="replace")
    except urllib.error.URLError as exc:
        raise RuntimeError(f"No pude escribir en el Sheet: {exc}") from exc
    normalized_body = body.replace(" ", "").lower()
    if '"ok":true' not in normalized_body:
        raise RuntimeError(f"El Sheet no confirmo guardado: {body[:300]}")
    if rows and '"escritos":0' in normalized_body:
        raise RuntimeError(
            "El Sheet respondio OK pero no escribio filas. "
            "Actualiza y redeploya el Apps Script con la ultima version de crear_sheet_planificacion.gs."
        )
    return "Google Sheet"


def save_daily_plan(selected_date: pd.Timestamp, focus_name: str, plan_df: pd.DataFrame) -> str:
    saved = load_saved_planner()
    new_rows = plan_df[["promotor", "PLANIFICADO"]].copy()
    new_rows["promotor"] = new_rows["promotor"].map(normalize_vendor_name)
    new_rows["planificado"] = pd.to_numeric(new_rows["PLANIFICADO"], errors="coerce")
    new_rows["fecha"] = pd.Timestamp(selected_date).normalize()
    new_rows["foco"] = focus_name
    new_rows = new_rows[["fecha", "foco", "promotor", "planificado"]]
    if not saved.empty:
        same_key = (saved["fecha"] == pd.Timestamp(selected_date).normalize()) & (saved["foco"] == focus_name)
        saved = saved.loc[~same_key].copy()
    output = pd.concat([saved, new_rows], ignore_index=True)
    local_path = local_planner_save(output)
    sheet_result = save_plan_to_google_sheet(selected_date, focus_name, plan_df)
    return f"{sheet_result} y respaldo local {local_path}" if sheet_result else local_path


def replace_saved_planner(saved: pd.DataFrame) -> str:
    saved = normalize_saved_planner(saved)
    return local_planner_save(saved)


def filter_focus(df: pd.DataFrame, focus_name: str) -> pd.DataFrame:
    result = df.copy()
    for column, allowed in PLANNER_FOCUS_RULES[focus_name].items():
        if column in {"title", "caption"}:
            continue
        if column in result.columns:
            result = result[result[column].astype(str).isin(allowed)]
    return result


def build_daily_planner_table(
    current_df: pd.DataFrame,
    selected_date: pd.Timestamp,
    focus_name: str,
    objectives: pd.DataFrame,
    saved_plan: pd.DataFrame,
) -> pd.DataFrame:
    focus_df = filter_focus(current_df, focus_name)
    real = (
        focus_df[focus_df["fecha"] == selected_date]
        .groupby(["mesa", "promotor"], as_index=False)["hl"]
        .sum()
        .rename(columns={"hl": "REAL"})
    )
    vendors = focus_df[["mesa", "promotor"]].drop_duplicates()
    if real.empty:
        real = vendors.assign(REAL=0.0)
    else:
        real = vendors.merge(real, on=["mesa", "promotor"], how="left").fillna({"REAL": 0.0})

    month_start = selected_date.replace(day=1)
    accum = (
        focus_df[(focus_df["fecha"] >= month_start) & (focus_df["fecha"] < selected_date)]
        .groupby("promotor", as_index=False)["hl"]
        .sum()
        .rename(columns={"hl": "ACUM. ANT."})
    )
    media_real = (
        focus_df[focus_df["fecha"] < selected_date]
        .groupby(["promotor", "fecha"], as_index=False)["hl"]
        .sum()
        .sort_values("fecha")
        .groupby("promotor", as_index=False)
        .tail(28)
        .groupby("promotor", as_index=False)["hl"]
        .mean()
        .rename(columns={"hl": "MEDIA REAL"})
    )
    table = real.merge(accum, on="promotor", how="left").merge(media_real, on="promotor", how="left")
    table["ACUM. ANT."] = table["ACUM. ANT."].fillna(0.0)

    if not objectives.empty:
        objective_lookup = objectives[objectives["foco"].map(normalize_planner_focus).eq(focus_name)].copy()
        objective_lookup["promotor_key"] = objective_lookup["promotor"].map(normalize_vendor_name)
        table["promotor_key"] = table["promotor"].map(normalize_vendor_name)
        table = table.merge(
            objective_lookup.drop_duplicates("promotor_key", keep="last")[["promotor_key", "objetivo_mes"]],
            on="promotor_key",
            how="left",
        )
    else:
        table["objetivo_mes"] = np.nan

    if not saved_plan.empty:
        plan = saved_plan[(saved_plan["fecha"] == selected_date) & (saved_plan["foco"] == focus_name)].copy()
        plan["promotor_key"] = plan["promotor"].map(normalize_vendor_name)
        if "promotor_key" not in table.columns:
            table["promotor_key"] = table["promotor"].map(normalize_vendor_name)
        plan_columns = ["promotor_key", "planificado"]
        for optional_column in ["hoja_origen", "celda_planificacion", "celda_objetivo"]:
            if optional_column in plan.columns:
                plan_columns.append(optional_column)
        table = table.merge(plan.drop_duplicates("promotor_key", keep="last")[plan_columns], on="promotor_key", how="left")
    else:
        table["planificado"] = np.nan
    for optional_column in ["hoja_origen", "celda_planificacion", "celda_objetivo"]:
        if optional_column not in table.columns:
            table[optional_column] = ""

    remaining_days = selling_days_remaining_from(selected_date)
    table["OBJETIVO MES"] = table["objetivo_mes"]
    table["DIAS HABILES MES"] = selling_days_in_month(selected_date)
    table["DIAS RESTANTES"] = remaining_days
    table["PLANIFICADO"] = table["planificado"].fillna(0.0)
    table["MEDIA NEC."] = np.where(
        (table["OBJETIVO MES"].fillna(0) != 0) & (remaining_days > 0),
        (table["OBJETIVO MES"] - table["ACUM. ANT."]) / remaining_days,
        np.nan,
    )
    table["MEDIA NEC."] = table["MEDIA NEC."].clip(lower=0)
    table["AVANCE"] = np.where(table["PLANIFICADO"].fillna(0) != 0, table["REAL"] / table["PLANIFICADO"] * 100, np.nan)
    table["VS MEDIA NEC."] = np.where(table["MEDIA NEC."].fillna(0) != 0, table["REAL"] / table["MEDIA NEC."] * 100, np.nan)
    table["VS MEDIA REAL"] = np.where(table["MEDIA REAL"].fillna(0) != 0, table["REAL"] / table["MEDIA REAL"] * 100, np.nan)
    columns = [
        "mesa",
        "promotor",
        "OBJETIVO MES",
        "ACUM. ANT.",
        "DIAS HABILES MES",
        "DIAS RESTANTES",
        "PLANIFICADO",
        "hoja_origen",
        "celda_planificacion",
        "REAL",
        "AVANCE",
        "MEDIA NEC.",
        "MEDIA REAL",
        "VS MEDIA NEC.",
        "VS MEDIA REAL",
    ]
    return table[columns].sort_values(["mesa", "REAL"], ascending=[True, False])


def metric_card(title: str, value: str, sub: str, accent: str = "blue") -> None:
    st.markdown(
        f"""
        <div class="metric-card accent-{accent}">
            <div class="metric-title">{title}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def business_card(title: str, value: float, total: float, class_name: str) -> None:
    share = (value / total * 100) if total else np.nan
    st.markdown(
        f"""
        <div class="business-card {class_name}">
            <div class="business-title">{title}</div>
            <div class="business-value">{format_hl(value)}</div>
            <div class="business-sub">{format_pct(share).replace("+", "")} del HL del dia seleccionado</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def chart_layout(fig):
    fig.update_layout(
        template="plotly_white",
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font=dict(family="Arial", color="#0f172a", size=13),
        title=dict(font=dict(color="#0f172a", size=18)),
        margin=dict(l=20, r=20, t=45, b=20),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(color="#0f172a"),
        ),
    )
    fig.update_xaxes(
        showgrid=False,
        color="#0f172a",
        title_font=dict(color="#0f172a"),
        tickfont=dict(color="#0f172a"),
    )
    fig.update_yaxes(
        gridcolor="rgba(16,24,40,.16)",
        color="#0f172a",
        title_font=dict(color="#0f172a"),
        tickfont=dict(color="#0f172a"),
    )
    fig.update_coloraxes(colorbar=dict(tickfont=dict(color="#0f172a"), title_font=dict(color="#0f172a")))
    return fig


def daily_business(df: pd.DataFrame, business: str) -> pd.DataFrame:
    business_df = df[df["unidad_negocio"] == business]
    if business_df.empty:
        return pd.DataFrame({"fecha": pd.to_datetime([]), "hl": pd.Series(dtype="float64")})
    return business_df.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")


def business_kpi_block(
    df: pd.DataFrame,
    selected_date: pd.Timestamp,
    business: str,
    accent: str,
    history_df: pd.DataFrame | None = None,
) -> None:
    daily = daily_business(df, business)
    history_daily = daily_business(history_df if history_df is not None else df, business)
    current = float(daily.loc[daily["fecha"] == selected_date, "hl"].sum()) if not daily.empty else 0.0
    stats = window_stats(history_daily, selected_date)[28] if not history_daily.empty else {
        "promedio": np.nan,
        "mediana": np.nan,
        "p25": np.nan,
        "p75": np.nan,
    }
    gap = current - stats["mediana"] if not pd.isna(stats["mediana"]) else np.nan
    planning_midpoint = (
        (stats["p25"] + stats["p75"]) / 2
        if not pd.isna(stats["p25"]) and not pd.isna(stats["p75"])
        else np.nan
    )

    st.markdown(f"#### {business}")
    cols = st.columns(6)
    cards = [
        ("HL dia actual", format_hl(current), selected_date.strftime("%d/%m/%Y"), accent),
        ("Promedio 28", format_hl(stats["promedio"]), "Dias habiles", "green"),
        ("Mediana 28", format_hl(stats["mediana"]), f"Gap {format_hl(gap)} HL", "violet"),
        ("Min planificable", format_hl(stats["p25"]), "Percentil 25", "orange"),
        ("Prom planificable", format_hl(planning_midpoint), "Entre min y max", "blue"),
        ("Max planificable", format_hl(stats["p75"]), "Percentil 75", "red"),
    ]
    for col, (title, value, sub, card_accent) in zip(cols, cards):
        with col:
            metric_card(title, value, sub, card_accent)

    compare_cols = st.columns(3)
    compare_cards = [
        (
            f"Vendido {months_back * 30} dias",
            format_hl(value),
            target_date.strftime("%d/%m/%Y"),
            card_accent,
        )
        for months_back, card_accent in zip(EXACT_MONTH_LOOKBACKS, (accent, "orange", "violet"))
        for target_date, value in [exact_previous_month_value(history_daily, selected_date, months_back)]
    ]
    for col, (title, value, sub, card_accent) in zip(compare_cols, compare_cards):
        with col:
            metric_card(title, value, sub, card_accent)


def planning_table(df: pd.DataFrame, selected_date: pd.Timestamp) -> pd.DataFrame:
    rows = []
    for group_name, group in df.groupby("calibre", dropna=False):
        daily = group.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")
        stats = window_stats(daily, selected_date)[28]
        rows.append(
            {
                "Calibre": group_name,
                "Minimo planificable": stats["p25"],
                "Mediana": stats["mediana"],
                "Promedio": stats["promedio"],
                "Maximo planificable": stats["p75"],
                "Min historico": stats["min"],
                "Max historico": stats["max"],
            }
        )
    table = pd.DataFrame(rows).sort_values("Promedio", ascending=False)
    return table


def promoter_planning_table(df: pd.DataFrame, selected_date: pd.Timestamp) -> pd.DataFrame:
    rows = []
    for promoter, group in df.groupby("promotor", dropna=False):
        daily = group.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")
        stats = window_stats(daily, selected_date)[28]
        current = float(daily.loc[daily["fecha"] == selected_date, "hl"].sum())
        planning_midpoint = (
            (stats["p25"] + stats["p75"]) / 2
            if not pd.isna(stats["p25"]) and not pd.isna(stats["p75"])
            else np.nan
        )
        exact_values = {
            f"Vendido {months_back * 30} dias": exact_previous_month_value(
                daily,
                selected_date,
                months_back,
            )[1]
            for months_back in EXACT_MONTH_LOOKBACKS
        }
        rows.append(
            {
                "Promotor": promoter,
                "HL dia actual": current,
                "Promedio 28": stats["promedio"],
                "Mediana 28": stats["mediana"],
                "Min planificable": stats["p25"],
                "Prom planificable": planning_midpoint,
                "Max planificable": stats["p75"],
                **exact_values,
            }
        )
    table = pd.DataFrame(rows)
    if table.empty:
        return table
    return table.sort_values("Promedio 28", ascending=False)


def year_comparison_curve(
    current_daily: pd.DataFrame,
    annual_daily: pd.DataFrame,
    selected_date: pd.Timestamp,
) -> pd.DataFrame:
    current_start = selected_date.replace(day=1)
    previous_start = current_start - pd.DateOffset(years=1)
    previous_end = selected_date - pd.DateOffset(years=1)

    current = current_daily[
        (current_daily["fecha"] >= current_start) & (current_daily["fecha"] <= selected_date)
    ].copy()
    current["serie"] = "Venta diaria actual"
    current["fecha_comparativa"] = current["fecha"]

    previous = annual_daily[
        (annual_daily["fecha"] >= previous_start) & (annual_daily["fecha"] <= previous_end)
    ].copy()
    previous["serie"] = "Venta diaria AA"
    previous["fecha_comparativa"] = previous["fecha"] + pd.DateOffset(years=1)

    return pd.concat([current, previous], ignore_index=True)


def aa_daily_report(current_df: pd.DataFrame, annual_df: pd.DataFrame) -> pd.DataFrame:
    dims = ["fecha", "division", "unidad_negocio", "calibre", "mesa", "canal", "supervisor", "promotor"]
    current = current_df.groupby(dims, as_index=False)["hl"].sum().rename(columns={"hl": "HL actual"})
    annual = annual_df.copy()
    annual["fecha"] = annual["fecha"] + pd.DateOffset(years=1)
    annual = annual.groupby(dims, as_index=False)["hl"].sum().rename(columns={"hl": "AA"})
    report = current.merge(annual, on=dims, how="left")
    report["Tendencia vs AA"] = np.where(
        report["AA"].fillna(0) != 0,
        (report["HL actual"] - report["AA"]) / report["AA"] * 100,
        np.nan,
    )
    return report.sort_values(["fecha", "HL actual"], ascending=[False, False])


SALES_CURVE_ORDER = ["CZA", "UNG", "MARKETPLACE", "VINO", "ADYACENCIAS"]
SALES_CURVE_COLORS = {
    "CZA": "#1463ff",
    "UNG": "#12b76a",
    "MARKETPLACE": "#f79009",
    "VINO": "#7a5af8",
    "ADYACENCIAS": "#f04438",
}


def sales_curve_group(value: object) -> str:
    text = str(value or "").strip().upper()
    if text in {"CZA", "UNG", "VINO", "ADYACENCIAS"}:
        return text
    if "MARKETPLACE" in text:
        return "MARKETPLACE"
    return ""


def daily_sales_curve_by_business(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "fecha" not in df.columns or "unidad_negocio" not in df.columns:
        return pd.DataFrame(columns=["fecha", "grupo_venta", "hl"])
    scoped = df.copy()
    scoped["grupo_venta"] = scoped["unidad_negocio"].map(sales_curve_group)
    scoped = scoped[scoped["grupo_venta"].isin(SALES_CURVE_ORDER)]
    if scoped.empty:
        return pd.DataFrame(columns=["fecha", "grupo_venta", "hl"])
    grouped = scoped.groupby(["fecha", "grupo_venta"], as_index=False)["hl"].sum()
    dates = pd.date_range(grouped["fecha"].min(), grouped["fecha"].max(), freq="D")
    full_index = pd.MultiIndex.from_product([dates, SALES_CURVE_ORDER], names=["fecha", "grupo_venta"])
    return (
        grouped.set_index(["fecha", "grupo_venta"])
        .reindex(full_index, fill_value=0.0)
        .reset_index()
    )


def executive_summary_table(
    current_df: pd.DataFrame,
    annual_df: pd.DataFrame | None,
    selected_date: pd.Timestamp,
    group_col: str,
    first_col: str,
    total_label: str | None = None,
    objectives_df: pd.DataFrame | None = None,
    objective_section: str = "",
) -> pd.DataFrame:
    current_start = selected_date.replace(day=1)
    previous_start = current_start - pd.DateOffset(years=1)
    previous_end = selected_date - pd.DateOffset(years=1)
    previous_month_end = previous_start + pd.offsets.MonthEnd(0)

    def aggregate(source: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp, value_name: str) -> pd.DataFrame:
        if source is None or source.empty or group_col not in source.columns:
            return pd.DataFrame(columns=[group_col, value_name])
        scoped = source[(source["fecha"] >= start) & (source["fecha"] <= end)]
        return scoped.groupby(group_col, as_index=False)["hl"].sum().rename(columns={"hl": value_name})

    accum_actual = aggregate(current_df, current_start, selected_date, "ACUM. ACTUAL")
    accum_ant = aggregate(current_df, current_start, selected_date - pd.Timedelta(days=1), "ACUM ANT.")
    today = aggregate(current_df, selected_date, selected_date, "HOY")
    aa = aggregate(annual_df, previous_start, previous_month_end, "AA")

    key_frames = [frame[[group_col]] for frame in (accum_actual, accum_ant, today, aa) if not frame.empty]
    if group_col == "canal":
        key_frames.append(pd.DataFrame({group_col: CANAL_ORDER}))
    if group_col == "division_informe":
        key_frames.append(pd.DataFrame({group_col: DIVISION_REPORT_ORDER}))
    keys = (
        pd.concat(key_frames, ignore_index=True).drop_duplicates()
        if key_frames
        else pd.DataFrame(columns=[group_col])
    )
    table = keys
    for frame in (accum_ant, accum_actual, today, aa):
        table = table.merge(frame, on=group_col, how="left")
    for column in ["ACUM ANT.", "ACUM. ACTUAL", "HOY", "AA"]:
        table[column] = table[column].fillna(0.0)
    table["TENDENCIA"] = table["ACUM. ACTUAL"].apply(lambda value: projected_month_trend(float(value), selected_date))
    table["TENDENCIA VS AA"] = np.nan
    valid_aa = table["AA"].fillna(0) != 0
    if valid_aa.any():
        table.loc[valid_aa, "TENDENCIA VS AA"] = (
            table.loc[valid_aa, "TENDENCIA"] / table.loc[valid_aa, "AA"] * 100
        )
    table = table.rename(columns={group_col: first_col})

    if objectives_df is not None and not objectives_df.empty:
        objective_lookup = objectives_df.copy()
        section_key = objective_section.strip().upper()
        if section_key:
            section_objectives = objective_lookup[objective_lookup["seccion"].isin(["", section_key])]
        else:
            section_objectives = objective_lookup
        section_objectives = section_objectives.drop_duplicates("item", keep="last")
        objective_map = section_objectives.set_index("item")["OBJ VTAS"]
        table["OBJ VTAS"] = table[first_col].astype(str).str.strip().str.upper().map(objective_map)
    else:
        table["OBJ VTAS"] = np.nan

    table["OBJ VS VENTAS"] = np.nan
    valid_objective = table["OBJ VTAS"].fillna(0) != 0
    if valid_objective.any():
        table.loc[valid_objective, "OBJ VS VENTAS"] = (
            table.loc[valid_objective, "ACUM. ACTUAL"] / table.loc[valid_objective, "OBJ VTAS"] * 100
        )
    table["TEND VS VENTAS"] = np.nan
    if valid_objective.any():
        table.loc[valid_objective, "TEND VS VENTAS"] = (
            table.loc[valid_objective, "TENDENCIA"] / table.loc[valid_objective, "OBJ VTAS"] * 100
        )

    ordered_cols = [
        first_col,
        "ACUM ANT.",
        "ACUM. ACTUAL",
        "HOY",
        "TENDENCIA",
        "AA",
        "TENDENCIA VS AA",
        "OBJ VTAS",
        "OBJ VS VENTAS",
        "TEND VS VENTAS",
    ]
    table = table[ordered_cols].sort_values("ACUM. ACTUAL", ascending=False)
    if group_col == "canal":
        order_map = {name: index for index, name in enumerate(CANAL_ORDER)}
        table["_orden"] = table[first_col].map(order_map).fillna(len(CANAL_ORDER))
        table = table.sort_values(["_orden", "ACUM. ACTUAL"], ascending=[True, False]).drop(columns=["_orden"])
    if group_col == "division_informe":
        table = table[table[first_col].astype(str).isin(DIVISION_REPORT_ORDER)].copy()
        order_map = {name: index for index, name in enumerate(DIVISION_REPORT_ORDER)}
        table["_orden"] = table[first_col].map(order_map).fillna(len(DIVISION_REPORT_ORDER))
        table = table.sort_values(["_orden", "ACUM. ACTUAL"], ascending=[True, False]).drop(columns=["_orden"])
        core_value_parts = ["CVZA CORE", "CVZA VALUE", "CVZA CORE +"]
        if "CVZA CORE + VALUE" in table[first_col].values:
            source_rows = table[table[first_col].isin(core_value_parts)]
            for column in ["ACUM ANT.", "ACUM. ACTUAL", "HOY", "TENDENCIA", "AA"]:
                table.loc[table[first_col] == "CVZA CORE + VALUE", column] = source_rows[column].sum()
        for label, business in (("TOTAL CVZA", "CZA"), ("TOTAL UNG", "UNG")):
            current_business = current_df[current_df["unidad_negocio"] == business] if "unidad_negocio" in current_df.columns else current_df.iloc[0:0]
            annual_business = (
                annual_df[annual_df["unidad_negocio"] == business]
                if annual_df is not None and "unidad_negocio" in annual_df.columns
                else None
            )
            total_table = executive_summary_table(
                current_business,
                annual_business,
                selected_date,
                "unidad_negocio",
                first_col,
                objectives_df=objectives_df,
                objective_section=objective_section,
            )
            if not total_table.empty:
                total_row = total_table.iloc[0].copy()
                total_row[first_col] = label
                if label in table[first_col].values:
                    for column in table.columns:
                        table.loc[table[first_col] == label, column] = total_row.get(column, np.nan)
        for pct_column in ["TENDENCIA VS AA", "OBJ VS VENTAS", "TEND VS VENTAS"]:
            table[pct_column] = np.nan
        valid_aa = table["AA"].fillna(0) != 0
        if valid_aa.any():
            table.loc[valid_aa, "TENDENCIA VS AA"] = (
                table.loc[valid_aa, "TENDENCIA"] / table.loc[valid_aa, "AA"] * 100
            )
        valid_objective = table["OBJ VTAS"].fillna(0) != 0
        if valid_objective.any():
            table.loc[valid_objective, "OBJ VS VENTAS"] = (
                table.loc[valid_objective, "ACUM. ACTUAL"] / table.loc[valid_objective, "OBJ VTAS"] * 100
            )
            table.loc[valid_objective, "TEND VS VENTAS"] = (
                table.loc[valid_objective, "TENDENCIA"] / table.loc[valid_objective, "OBJ VTAS"] * 100
            )

    if total_label:
        if group_col == "division_informe":
            report_current_df = (
                current_df[current_df["unidad_negocio"].isin(REPORT_TOTAL_UNITS)]
                if "unidad_negocio" in current_df.columns
                else current_df
            )
            report_annual_df = (
                annual_df[annual_df["unidad_negocio"].isin(REPORT_TOTAL_UNITS)]
                if annual_df is not None and not annual_df.empty and "unidad_negocio" in annual_df.columns
                else annual_df
            )
            total_accum_actual = report_current_df.loc[
                (report_current_df["fecha"] >= current_start) & (report_current_df["fecha"] <= selected_date), "hl"
            ].sum()
            total_accum_ant = report_current_df.loc[
                (report_current_df["fecha"] >= current_start) & (report_current_df["fecha"] <= selected_date - pd.Timedelta(days=1)), "hl"
            ].sum()
            total_today = report_current_df.loc[report_current_df["fecha"] == selected_date, "hl"].sum()
            total_aa = (
                report_annual_df.loc[(report_annual_df["fecha"] >= previous_start) & (report_annual_df["fecha"] <= previous_month_end), "hl"].sum()
                if report_annual_df is not None and not report_annual_df.empty
                else 0.0
            )
            total_obj = table.loc[table[first_col].isin(["TOTAL CVZA", "TOTAL UNG", "AGUAS ECO", "VINO", "ADYACENCIAS"]), "OBJ VTAS"].sum(min_count=1)
            total = {
                first_col: total_label,
                "ACUM ANT.": total_accum_ant,
                "ACUM. ACTUAL": total_accum_actual,
                "HOY": total_today,
                "TENDENCIA": projected_month_trend(float(total_accum_actual), selected_date),
                "AA": total_aa,
                "OBJ VTAS": total_obj,
            }
        else:
            total = {
                first_col: total_label,
                "ACUM ANT.": table["ACUM ANT."].sum(),
                "ACUM. ACTUAL": table["ACUM. ACTUAL"].sum(),
                "HOY": table["HOY"].sum(),
                "TENDENCIA": table["TENDENCIA"].sum(),
                "AA": table["AA"].sum(),
                "OBJ VTAS": table["OBJ VTAS"].sum(min_count=1),
            }
        total["TENDENCIA VS AA"] = total["TENDENCIA"] / total["AA"] * 100 if total["AA"] else np.nan
        total["OBJ VS VENTAS"] = (
            total["ACUM. ACTUAL"] / total["OBJ VTAS"] * 100
            if valid_nonzero(total["OBJ VTAS"])
            else np.nan
        )
        total["TEND VS VENTAS"] = (
            total["TENDENCIA"] / total["OBJ VTAS"] * 100
            if valid_nonzero(total["OBJ VTAS"])
            else np.nan
        )
        table = pd.concat([pd.DataFrame([total]), table], ignore_index=True)
    return table


def format_exec_number(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "-"
    if abs(float(value)) >= 100:
        text = f"{value:,.0f}"
    else:
        text = f"{value:,.1f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def render_exec_table(title: str, table: pd.DataFrame, first_col: str) -> None:
    rows = []
    for idx, row in table.iterrows():
        row_class = " class=\"total-row\"" if idx == 0 and str(row[first_col]).upper().startswith(("TOTAL", "HL TOTAL")) else ""
        cells = [f"<td>{row[first_col]}</td>"]
        for column in table.columns:
            if column == first_col:
                continue
            if column in {"TENDENCIA VS AA", "OBJ VS VENTAS", "TEND VS VENTAS"}:
                cells.append(f"<td>{format_pct(row[column]).replace('+', '')}</td>")
            else:
                cells.append(f"<td>{format_exec_number(row[column])}</td>")
        rows.append(f"<tr{row_class}>{''.join(cells)}</tr>")
    header = "".join(f"<th>{column}</th>" for column in table.columns)
    st.markdown(
        f"""
        <div class="exec-wrap">
            <div class="exec-title">{title}</div>
            <table class="exec-table">
                <thead><tr>{header}</tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


def pct_class(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return "good" if float(value) >= 100 else "bad"


def render_planner_table(title: str, focus_name: str, table: pd.DataFrame) -> None:
    headers = ["", "PLANIFICADO", "REAL", "AVANCE", "MEDIA NEC.", "MEDIA REAL", "VS MEDIA NEC.", "VS MEDIA REAL"]
    percent_cols = {"AVANCE", "VS MEDIA NEC.", "VS MEDIA REAL"}
    yellow_headers = {"PLANIFICADO", "VS MEDIA NEC.", "VS MEDIA REAL"}
    rows: list[str] = []
    total = {
        "PLANIFICADO": table["PLANIFICADO"].sum(min_count=1),
        "REAL": table["REAL"].sum(min_count=1),
        "MEDIA NEC.": table["MEDIA NEC."].sum(min_count=1),
        "MEDIA REAL": table["MEDIA REAL"].sum(min_count=1),
    }
    total["AVANCE"] = total["REAL"] / total["PLANIFICADO"] * 100 if valid_nonzero(total["PLANIFICADO"]) else np.nan
    total["VS MEDIA NEC."] = total["REAL"] / total["MEDIA NEC."] * 100 if valid_nonzero(total["MEDIA NEC."]) else np.nan
    total["VS MEDIA REAL"] = total["REAL"] / total["MEDIA REAL"] * 100 if valid_nonzero(total["MEDIA REAL"]) else np.nan
    total_cells = ["<td>TOTAL</td>"]
    for column in headers[1:]:
        cls = pct_class(total[column]) if column in percent_cols else ""
        total_cells.append(f"<td class='{cls}'>{format_pct_plain(total[column]) if column in percent_cols else format_hl(total[column])}</td>")
    rows.append(f"<tr class='total-row'>{''.join(total_cells)}</tr>")

    for mesa, group in table.groupby("mesa", dropna=False):
        rows.append(f"<tr class='mesa-row'><td colspan='8'>{mesa}</td></tr>")
        for _, row in group.iterrows():
            cells = [f"<td>{row['promotor']}</td>"]
            for column in headers[1:]:
                cls = pct_class(row[column]) if column in percent_cols else ""
                cells.append(f"<td class='{cls}'>{format_pct_plain(row[column]) if column in percent_cols else format_hl(row[column])}</td>")
            rows.append(f"<tr>{''.join(cells)}</tr>")

    header_cells = "".join(f"<th class='{'yellow' if header in yellow_headers else ''}'>{header}</th>" for header in headers)
    caption = PLANNER_FOCUS_RULES[focus_name]["caption"]
    st.markdown(
        f"""
        <div class="exec-wrap">
            <div class="exec-title">{title}</div>
            <div class="planner-note">{focus_name}<br>{caption}</div>
            <table class="planner-table">
                <thead><tr>{header_cells}</tr></thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


def default_objectives() -> pd.DataFrame:
    return pd.DataFrame(DEFAULT_OBJECTIVES_ROWS)


def make_pdf_cell(value, is_percent: bool = False) -> str:
    if is_percent:
        return format_pct(value).replace("+", "")
    return format_exec_number(value)


def table_for_pdf(table: pd.DataFrame, first_col: str) -> list[list[str]]:
    output = [list(table.columns)]
    percent_cols = {"TENDENCIA VS AA", "OBJ VS VENTAS", "TEND VS VENTAS"}
    for _, row in table.iterrows():
        output_row = [str(row[first_col])]
        for column in table.columns:
            if column == first_col:
                continue
            output_row.append(make_pdf_cell(row[column], column in percent_cols))
        output.append(output_row)
    return output


def build_report_pdf(tables: list[tuple[str, pd.DataFrame, str]], selected_date: pd.Timestamp) -> bytes:
    buffer = io.BytesIO()
    page_width, _ = landscape(A4)
    margin = 18
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=margin,
        leftMargin=margin,
        topMargin=18,
        bottomMargin=18,
        title="Informe venta del dia",
    )
    usable_width = page_width - 2 * margin
    header_color = colors.HexColor("#28549A")
    first_col_color = colors.HexColor("#2F5EA8")
    title_style = ParagraphStyle(
        "ReportTitle",
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#111827"),
        spaceBefore=6,
        spaceAfter=4,
    )
    elements = [Paragraph(f"Informe venta del dia - {selected_date.strftime('%d/%m/%Y')}", title_style)]

    compact_cols = {"ACUM ANT.", "ACUM. ACTUAL", "HOY", "TENDENCIA", "AA", "TENDENCIA VS AA"}
    full_cols = [
        None,
        "ACUM ANT.",
        "ACUM. ACTUAL",
        "HOY",
        "TENDENCIA",
        "AA",
        "TENDENCIA VS AA",
        "OBJ VTAS",
        "OBJ VS VENTAS",
        "TEND VS VENTAS",
    ]

    for index, (title, table, first_col) in enumerate(tables):
        if table.empty:
            continue
        if index in {2, 3, 5, 6}:
            table_to_print = table[[first_col] + [col for col in table.columns if col in compact_cols]].copy()
        else:
            table_to_print = table.copy()
        elements.append(Paragraph(title.upper(), section_style))
        pdf_table = table_for_pdf(table_to_print, first_col)
        col_count = len(pdf_table[0])
        first_width = min(128, usable_width * 0.26)
        rest_width = (usable_width - first_width) / max(col_count - 1, 1)
        report_table = Table(pdf_table, colWidths=[first_width] + [rest_width] * (col_count - 1), repeatRows=1)
        style = TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 6.8),
                ("FONTSIZE", (0, 1), (-1, -1), 6.6),
                ("BACKGROUND", (0, 0), (-1, 0), header_color),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (0, -1), first_col_color),
                ("TEXTCOLOR", (0, 1), (0, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
        report_table.setStyle(style)
        elements.append(report_table)
        elements.append(Spacer(1, 8))
        if index in {3}:
            elements.append(PageBreak())

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def planner_table_for_pdf(table: pd.DataFrame) -> list[list[str]]:
    columns = ["promotor", "OBJETIVO MES", "ACUM. ANT.", "PLANIFICADO", "REAL", "AVANCE", "MEDIA NEC.", "MEDIA REAL", "VS MEDIA NEC.", "VS MEDIA REAL"]
    headers = ["VENDEDOR", "OBJ MES", "ACUM ANT", "PLANIF", "REAL", "AVANCE", "MEDIA NEC", "MEDIA REAL", "VS M NEC", "VS M REAL"]
    output = [headers]
    for _, row in table[columns].iterrows():
        output.append(
            [
                str(row["promotor"]),
                format_hl(row["OBJETIVO MES"]),
                format_hl(row["ACUM. ANT."]),
                format_hl(row["PLANIFICADO"]),
                format_hl(row["REAL"]),
                format_pct_plain(row["AVANCE"]),
                format_hl(row["MEDIA NEC."]),
                format_hl(row["MEDIA REAL"]),
                format_pct_plain(row["VS MEDIA NEC."]),
                format_pct_plain(row["VS MEDIA REAL"]),
            ]
        )
    return output


def build_planner_pdf(tables: list[tuple[str, str, pd.DataFrame]], selected_date: pd.Timestamp) -> bytes:
    buffer = io.BytesIO()
    page_width, _ = landscape(A4)
    margin = 16
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=margin,
        leftMargin=margin,
        topMargin=16,
        bottomMargin=16,
        title="Cierre planificador diario",
    )
    usable_width = page_width - 2 * margin
    elements = [
        Paragraph(
            f"Cierre planificador diario - {selected_date.strftime('%d/%m/%Y')}",
            ParagraphStyle("PlannerTitle", fontName="Helvetica-Bold", fontSize=16, leading=18, textColor=colors.HexColor("#111827"), spaceAfter=8),
        )
    ]
    col_widths = [usable_width * 0.22] + [usable_width * 0.078] * 9
    for index, (title, caption, table) in enumerate(tables):
        if table.empty:
            continue
        elements.append(
            Paragraph(
                f"{title} - {caption}",
                ParagraphStyle("PlannerSection", fontName="Helvetica-Bold", fontSize=9, leading=11, textColor=colors.HexColor("#111827"), spaceBefore=5, spaceAfter=3),
            )
        )
        report_table = Table(planner_table_for_pdf(table), colWidths=col_widths, repeatRows=1)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0070C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 6.2),
            ("FONTSIZE", (0, 1), (-1, -1), 6.0),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]
        for row_index, (_, row) in enumerate(table.iterrows(), start=1):
            for col_index, column in ((5, "AVANCE"), (8, "VS MEDIA NEC."), (9, "VS MEDIA REAL")):
                if pd.isna(row[column]):
                    continue
                style_commands.append(("BACKGROUND", (col_index, row_index), (col_index, row_index), colors.HexColor("#C6EFCE") if row[column] >= 100 else colors.HexColor("#FFC7CE")))
        report_table.setStyle(TableStyle(style_commands))
        elements.extend([report_table, Spacer(1, 7)])
        if index == 1:
            elements.append(PageBreak())
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def main() -> None:
    page_setup()
    st.markdown(
        """
        <div class="hero">
            <h1>Venta diaria HL</h1>
            <p>Analisis comercial con ventanas habiles, planificacion por percentiles y comparacion mensual.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.sidebar.title("Datos y filtros")
    st.sidebar.link_button("Ir al planificador", secret_or_env("SMALL_DASH_URL", SMALL_DASH_URL), width="stretch")
    if "force_data_refresh" not in st.session_state:
        st.session_state["force_data_refresh"] = False
    if st.sidebar.button("Actualizar datos", width="stretch"):
        st.session_state["force_data_refresh"] = True
        st.cache_data.clear()
        st.rerun()
    data_dir = current_data_dir(force_refresh=bool(st.session_state.get("force_data_refresh", False)))
    st.session_state["force_data_refresh"] = False
    st.sidebar.caption(f"Carpeta automatica: {data_dir}")
    load_annual_comparison = st.sidebar.checkbox(
        "Cargar comparacion AA",
        value=True,
        help="Activalo cuando necesites comparar contra venta anual. El archivo anual es pesado y puede demorar.",
    )

    uploaded = st.sidebar.file_uploader("Carga manual si falla la carpeta", type=["txt", "csv"])

    try:
        if uploaded is not None:
            df, info = load_source_from_upload(uploaded.name, uploaded.getvalue())
        else:
            latest = latest_daily_file_in_folder(data_dir)
            if latest is None:
                st.warning("No encontre archivos TXT/CSV en la carpeta automatica. Use la carga manual.")
                st.stop()
            df, info = load_source_from_path(str(latest), latest.stat().st_mtime_ns)
    except Exception as exc:
        st.error(f"No pude leer el archivo: {exc}")
        st.stop()

    if df.empty:
        st.warning("El archivo no contiene filas validas luego de normalizar y excluir domingos.")
        st.stop()

    customer_channels: pd.DataFrame | None = None
    customer_info: SourceInfo | None = None
    customer_file = latest_customer_file_in_folder(data_dir)
    if customer_file is not None:
        try:
            customer_channels, customer_info = load_customer_channels(str(customer_file), customer_file.stat().st_mtime_ns)
            df = apply_customer_channels(df, customer_channels)
        except Exception as exc:
            st.sidebar.warning(f"No pude leer la planilla de clientes para canal: {exc}")
    else:
        st.sidebar.warning("No se encontro planilla de clientes para clasificar canal.")
    df = ensure_analysis_columns(df)

    aux_segments: dict[str, pd.DataFrame] | None = None
    aux_info: SourceInfo | None = None
    aux_file = latest_auxiliary_file_in_folder(data_dir)
    if aux_file is not None:
        try:
            aux_segments, aux_info = load_auxiliary_segments(str(aux_file), aux_file.stat().st_mtime_ns)
            df = apply_auxiliary_segments(df, aux_segments)
        except Exception as exc:
            st.sidebar.warning(f"No pude leer auxiliares para segmentos: {exc}")
    else:
        df = apply_auxiliary_segments(df, None)
        st.sidebar.warning("No se encontro archivo auxiliares para segmentos.")
    df = ensure_analysis_columns(df)

    objectives_df = default_objectives()
    objectives_info: SourceInfo | None = None
    objectives_file = latest_objectives_file_in_folder(data_dir)
    if objectives_file is not None:
        try:
            objectives_df, objectives_info = load_objectives(str(objectives_file), objectives_file.stat().st_mtime_ns)
        except Exception as exc:
            st.sidebar.warning(f"No pude leer objetivos; uso objetivos de mayo de respaldo: {exc}")

    annual_df: pd.DataFrame | None = None
    annual_info: SourceInfo | None = None
    annual_warning = "No se encontro archivo de venta anual para comparacion AA"
    if uploaded is not None:
        st.sidebar.info("La comparacion AA automatica se omite cuando se usa carga manual.")
    elif load_annual_comparison:
        annual_file = latest_annual_file_in_folder(data_dir)
        if annual_file is not None:
            try:
                annual_df, annual_info = load_annual_source_from_path(str(annual_file), annual_file.stat().st_mtime_ns)
                annual_df = apply_customer_channels(annual_df, customer_channels)
                annual_df = apply_auxiliary_segments(annual_df, aux_segments)
                annual_df = ensure_analysis_columns(annual_df)
            except Exception as exc:
                st.sidebar.warning(f"{annual_warning}: {exc}")
    else:
        st.sidebar.info("AA desactivado para abrir rapido. Marque 'Cargar comparacion AA' si lo necesita.")

    st.sidebar.success(f"Fuente: {info.label}")
    if info.modified:
        st.sidebar.caption(f"Modificado: {info.modified}")
    if info.path:
        st.sidebar.caption(info.path)
    if customer_info is not None:
        st.sidebar.success(f"Clientes: {customer_info.label}")
        if customer_info.modified:
            st.sidebar.caption(f"Modificado clientes: {customer_info.modified}")
    if objectives_info is not None:
        st.sidebar.success(f"Objetivos: {objectives_info.label}")
    else:
        st.sidebar.info("Objetivos: respaldo mayo cargado hasta que agregues objetivos.xlsx")
    if aux_info is not None:
        st.sidebar.success(f"Auxiliares: {aux_info.label}")
    if annual_info is not None:
        st.sidebar.success(f"AA: {annual_info.label}")
        if annual_info.modified:
            st.sidebar.caption(f"Modificado AA: {annual_info.modified}")
    elif load_annual_comparison:
        st.sidebar.warning(annual_warning)

    filtered, selected_date, dimension_filters = apply_filters(df)
    annual_filtered = apply_dimension_filters(ensure_analysis_columns(annual_df), dimension_filters) if annual_df is not None else None
    if filtered.empty:
        st.warning("No hay datos para los filtros seleccionados.")
        st.stop()

    historical_filtered = combine_current_with_history(filtered, annual_filtered)
    daily = filtered.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")
    historical_daily = historical_filtered.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")
    annual_daily = (
        annual_filtered.groupby("fecha", as_index=False)["hl"].sum().sort_values("fecha")
        if annual_filtered is not None and not annual_filtered.empty
        else pd.DataFrame({"fecha": pd.to_datetime([]), "hl": pd.Series(dtype="float64")})
    )
    current_value = float(daily.loc[daily["fecha"] == selected_date, "hl"].sum())
    stats = window_stats(historical_daily, selected_date)
    baseline = stats[28]
    previous_same = previous_month_same_business_day(historical_daily, selected_date)
    variation = ((current_value - previous_same) / previous_same * 100) if previous_same else np.nan
    gap_median = current_value - baseline["mediana"]
    current_by_business = (
        filtered.loc[filtered["fecha"] == selected_date]
        .groupby("unidad_negocio", as_index=True)["hl"]
        .sum()
    )

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        metric_card("HL dia actual", format_hl(current_value), selected_date.strftime("%d/%m/%Y"), "blue")
    with k2:
        metric_card("Mediana 28 habiles", format_hl(baseline["mediana"]), f"Gap {format_hl(gap_median)} HL", "violet")
    with k3:
        metric_card("Promedio 28 habiles", format_hl(baseline["promedio"]), "Base historica filtrada", "green")
    with k4:
        metric_card("Min planificable", format_hl(baseline["p25"]), "Percentil 25", "orange")
    with k5:
        metric_card("Max planificable", format_hl(baseline["p75"]), f"Var mes ant. {format_pct(variation)}", "red")

    if annual_info is not None and not annual_daily.empty:
        aa_date, aa_value = exact_previous_year_value(annual_daily, selected_date)
        aa_accum = accumulated_vs_previous_year(daily, annual_daily, selected_date)
        st.subheader("Comparacion AA")
        aa1, aa2, aa3, aa4 = st.columns(4)
        with aa1:
            metric_card("AA dia exacto", format_hl(aa_value), aa_date.strftime("%d/%m/%Y"), "violet")
        with aa2:
            metric_card("Acumulado actual", format_hl(aa_accum["current_value"]), "Periodo seleccionado", "blue")
        with aa3:
            metric_card("Acumulado AA", format_hl(aa_accum["previous_value"]), "Mismo periodo AA", "orange")
        with aa4:
            metric_card("Tendencia vs AA", format_pct(aa_accum["trend"]), f"Proy {format_hl(aa_accum['projected_value'])} HL", "green")

    st.subheader("Venta del mismo dia en meses anteriores")
    exact_cols = st.columns(3)
    for col, months_back, accent in zip(exact_cols, EXACT_MONTH_LOOKBACKS, ("blue", "orange", "violet")):
        target_date, value = exact_previous_month_value(historical_daily, selected_date, months_back)
        with col:
            metric_card(
                f"Vendido {months_back * 30} dias",
                format_hl(value),
                target_date.strftime("%d/%m/%Y"),
                accent,
            )

    st.subheader("Separado por negocio")
    b1, b2 = st.columns(2)
    with b1:
        business_card("CZA - HL dia actual", float(current_by_business.get("CZA", 0.0)), current_value, "business-cza")
    with b2:
        business_card("UNG - HL dia actual", float(current_by_business.get("UNG", 0.0)), current_value, "business-ung")

    st.subheader("KPIs completos por negocio")
    business_kpi_block(filtered, selected_date, "CZA", "blue", historical_filtered)
    business_kpi_block(filtered, selected_date, "UNG", "green", historical_filtered)

    st.subheader("Ventanas habiles")
    cols = st.columns(4)
    for col, window in zip(cols, WINDOWS):
        with col:
            metric_card(
                f"{window} dias",
                format_hl(stats[window]["promedio"]),
                f"Mediana {format_hl(stats[window]['mediana'])}",
                "blue" if window == 7 else "green" if window == 14 else "orange" if window == 21 else "violet",
            )

    tab_overview, tab_informe, tab_aa, tab_rankings, tab_promoters, tab_planning, tab_base = st.tabs(
        ["Evolucion", "Informe", "AA", "Rankings", "Promotores", "Planificacion", "Base normalizada"]
    )

    with tab_overview:
        sales_curve = daily_sales_curve_by_business(filtered)
        if sales_curve.empty:
            line = px.line(
                daily,
                x="fecha",
                y="hl",
                markers=True,
                title="HL diarios",
                color_discrete_sequence=["#1463ff"],
            )
        else:
            line = px.line(
                sales_curve,
                x="fecha",
                y="hl",
                color="grupo_venta",
                markers=True,
                title="HL diarios por negocio",
                category_orders={"grupo_venta": SALES_CURVE_ORDER},
                color_discrete_map=SALES_CURVE_COLORS,
                labels={"grupo_venta": "Negocio"},
            )
            line.update_xaxes(title_text="Fecha")
        line.update_yaxes(title_text="HL")
        st.plotly_chart(chart_layout(line), width="stretch")

        left, right = st.columns(2)
        by_calibre = filtered.groupby("calibre", as_index=False)["hl"].sum().sort_values("hl", ascending=False).head(20)
        by_business = filtered.groupby("unidad_negocio", as_index=False)["hl"].sum().sort_values("hl", ascending=False)
        with left:
            fig = px.bar(
                by_calibre,
                x="hl",
                y="calibre",
                orientation="h",
                title="HL por calibre",
                color="hl",
                color_continuous_scale=["#00a7c8", "#1463ff", "#7a5af8"],
            )
            st.plotly_chart(chart_layout(fig), width="stretch")
        with right:
            fig = px.bar(
                by_business,
                x="unidad_negocio",
                y="hl",
                title="HL por negocio CZA/UNG",
                color="unidad_negocio",
                color_discrete_sequence=["#1463ff", "#12b76a", "#f79009"],
            )
            st.plotly_chart(chart_layout(fig), width="stretch")

    with tab_informe:
        st.subheader("Informe de venta del dia")
        if annual_info is None or annual_filtered is None or annual_filtered.empty:
            st.warning("No se encontro archivo de venta anual para comparacion AA")

        report_source = annual_filtered if annual_filtered is not None and not annual_filtered.empty else None
        cerveza_df = filtered[filtered["unidad_negocio"] == "CZA"]
        cerveza_aa = report_source[report_source["unidad_negocio"] == "CZA"] if report_source is not None else None
        ung_df = filtered[filtered["unidad_negocio"] == "UNG"]
        ung_aa = report_source[report_source["unidad_negocio"] == "UNG"] if report_source is not None else None

        informe_tables = [
            (
                "Division / negocio",
                executive_summary_table(
                    filtered,
                    report_source,
                    selected_date,
                    "division_informe",
                    "DIVISION",
                    "HL TOTALES",
                    objectives_df,
                    "DIVISION",
                ),
                "DIVISION",
            ),
            (
                "X mesa - cerveza",
                executive_summary_table(cerveza_df, cerveza_aa, selected_date, "mesa", "MESA", objectives_df=objectives_df, objective_section="MESA CERVEZA"),
                "MESA",
            ),
            (
                "X calibre - cerveza",
                executive_summary_table(cerveza_df, cerveza_aa, selected_date, "calibre", "CALIBRE", objectives_df=objectives_df, objective_section="CALIBRE CERVEZA"),
                "CALIBRE",
            ),
            (
                "X canal - cerveza",
                executive_summary_table(cerveza_df, cerveza_aa, selected_date, "canal", "CANAL", objectives_df=objectives_df, objective_section="CANAL CERVEZA"),
                "CANAL",
            ),
            (
                "X mesa - ung",
                executive_summary_table(ung_df, ung_aa, selected_date, "mesa", "MESA", objectives_df=objectives_df, objective_section="MESA UNG"),
                "MESA",
            ),
            (
                "X calibre - ung",
                executive_summary_table(ung_df, ung_aa, selected_date, "calibre", "CALIBRE", objectives_df=objectives_df, objective_section="CALIBRE UNG"),
                "CALIBRE",
            ),
            (
                "X canal - ung",
                executive_summary_table(ung_df, ung_aa, selected_date, "canal", "CANAL", objectives_df=objectives_df, objective_section="CANAL UNG"),
                "CANAL",
            ),
        ]

        pdf_bytes = build_report_pdf(informe_tables, selected_date)
        st.download_button(
            "Exportar informe PDF",
            data=pdf_bytes,
            file_name=f"informe_venta_{selected_date.strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            width="stretch",
        )

        render_exec_table(*informe_tables[0])

        left, right = st.columns(2)
        with left:
            render_exec_table(*informe_tables[1])
            render_exec_table(*informe_tables[2])
            render_exec_table(*informe_tables[3])
        with right:
            render_exec_table(*informe_tables[4])
            render_exec_table(*informe_tables[5])
            render_exec_table(*informe_tables[6])

    with tab_aa:
        if annual_info is None or annual_filtered is None or annual_filtered.empty:
            st.warning("No se encontro archivo de venta anual para comparacion AA")
        else:
            curve = year_comparison_curve(daily, annual_daily, selected_date)
            if curve.empty:
                st.info("No hay datos AA para el periodo y filtros seleccionados.")
            else:
                fig = px.line(
                    curve,
                    x="fecha_comparativa",
                    y="hl",
                    color="serie",
                    markers=True,
                    title="Curva comparativa: venta diaria actual vs AA",
                    color_discrete_map={
                        "Venta diaria actual": "#1463ff",
                        "Venta diaria AA": "#f79009",
                    },
                )
                fig.update_xaxes(title_text="Fecha comparable")
                fig.update_yaxes(title_text="HL")
                st.plotly_chart(chart_layout(fig), width="stretch")

            aa_table = annual_filtered.copy()
            aa_table["fecha_comparativa"] = aa_table["fecha"] + pd.DateOffset(years=1)
            report = aa_daily_report(filtered, annual_filtered)
            st.subheader("Informe diario con columna AA")
            st.dataframe(
                report.style.format(
                    {
                        "HL actual": format_hl,
                        "AA": format_hl,
                        "Tendencia vs AA": format_pct,
                    }
                ),
                width="stretch",
                hide_index=True,
            )

            st.subheader("Base historica AA filtrada")
            st.dataframe(
                aa_table[
                    [
                        "fecha",
                        "fecha_comparativa",
                        "division",
                        "unidad_negocio",
                        "calibre",
                        "mesa",
                        "canal",
                        "supervisor",
                        "promotor",
                        "hl",
                    ]
                ].sort_values("fecha", ascending=False),
                width="stretch",
                hide_index=True,
            )

    with tab_rankings:
        left, right = st.columns(2)
        for container, column, title, colors in [
            (left, "supervisor", "Ranking de supervisores", ["#12b76a", "#1463ff"]),
            (right, "promotor", "Ranking de promotores", ["#f79009", "#f04438"]),
        ]:
            ranking = filtered.groupby(column, as_index=False)["hl"].sum().sort_values("hl", ascending=False).head(15)
            fig = px.bar(
                ranking,
                x="hl",
                y=column,
                orientation="h",
                title=title,
                color="hl",
                color_continuous_scale=colors,
            )
            container.plotly_chart(chart_layout(fig), width="stretch")

    with tab_promoters:
        promoter_table = promoter_planning_table(filtered, selected_date)
        st.dataframe(
            promoter_table.style.format(
                {
                    "HL dia actual": format_hl,
                    "Promedio 28": format_hl,
                    "Mediana 28": format_hl,
                    "Min planificable": format_hl,
                    "Prom planificable": format_hl,
                    "Max planificable": format_hl,
                    "Vendido 30 dias": format_hl,
                    "Vendido 60 dias": format_hl,
                    "Vendido 90 dias": format_hl,
                }
            ),
            width="stretch",
            hide_index=True,
        )

    with tab_planning:
        table = planning_table(filtered, selected_date)
        st.dataframe(
            table.style.format(
                {
                    "Minimo planificable": format_hl,
                    "Mediana": format_hl,
                    "Promedio": format_hl,
                    "Maximo planificable": format_hl,
                    "Min historico": format_hl,
                    "Max historico": format_hl,
                }
            ),
            width="stretch",
            hide_index=True,
        )

    with tab_base:
        st.caption(f"{len(filtered):,} filas filtradas de {len(df):,} filas normalizadas.".replace(",", "."))
        st.dataframe(filtered.sort_values("fecha", ascending=False), width="stretch", hide_index=True)


if __name__ == "__main__":
    main()
